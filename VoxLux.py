#!/usr/bin/env python3
"""VoxLux V3 — Transcription + Aperçu sous-titres + Traduction + Export SRT/VTT/Excel"""

import os, re, json, platform, tempfile, subprocess, threading, webbrowser, uuid
import urllib.request, urllib.error
from pathlib import Path
from datetime import datetime
import numpy as np
from flask import Flask, request, jsonify, send_file, Response
import whisper
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024 * 1024  # 4 GB

UPLOAD_DIR = tempfile.mkdtemp(prefix='voxlux_')

# ─── Mise à jour automatique (GitHub Releases) ───────────────────────────────
APP_VERSION  = "1.0.1"
GITHUB_REPO  = "luxmodernis/VoxLux"

def _version_tuple(v):
    v = (v or '').strip().lstrip('vV')
    parts = []
    for p in v.split('.'):
        try: parts.append(int(p))
        except ValueError: parts.append(0)
    return tuple(parts) or (0,)

# ─── Projets — vrais fichiers sur disque (survit à un crash/redémarrage) ─────
# VOXLUX_PROJECTS_DIR permet de rediriger ce dossier (utilisé pour les tests
# de développement) — ne JAMAIS lancer de serveur de test sans cette variable
# positionnée sur un chemin isolé : le défaut touche le vrai Documents de l'utilisateur.
PROJECTS_DIR = Path(os.environ.get('VOXLUX_PROJECTS_DIR') or (Path.home() / "Documents" / "VoxLux Projets"))
PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_SUFFIX = '.voxluxproj.json'

def safe_project_name(name):
    name = (name or '').strip()
    name = re.sub(r'[\\/:*?"<>|]', '', name)   # caractères interdits dans un nom de fichier
    name = re.sub(r'\s+', ' ', name).strip()
    return name[:120] if name else 'Projet sans titre'

def project_display_name(path):
    n = path.name
    return n[:-len(PROJECT_SUFFIX)] if n.endswith(PROJECT_SUFFIX) else path.stem

print("⏳ Chargement du modèle Whisper…")
model = whisper.load_model("small")
print("✅ Modèle prêt.")

# ─── Utilitaires audio/vidéo ──────────────────────────────────────────────────

def find_ffmpeg():
    """Trouve ffmpeg : binaire système d'abord, sinon celui fourni par imageio-ffmpeg (pip).
    Fonctionne sur macOS, Windows et Linux sans installation système."""
    for cmd in ["ffmpeg", "/usr/local/bin/ffmpeg", "/opt/homebrew/bin/ffmpeg"]:
        try:
            subprocess.run([cmd, "-version"], capture_output=True, check=True)
            return cmd
        except Exception:
            continue
    try:
        import imageio_ffmpeg
        exe = imageio_ffmpeg.get_ffmpeg_exe()
        subprocess.run([exe, "-version"], capture_output=True, check=True)
        return exe
    except Exception:
        return None

FFMPEG = find_ffmpeg()

# Évite l'ouverture d'une fenêtre console sur Windows lors des appels ffmpeg
_NO_WINDOW = {'creationflags': subprocess.CREATE_NO_WINDOW} if os.name == 'nt' else {}

def load_audio(path, sr=16000):
    """Décode n'importe quel fichier audio/vidéo en tableau float32 mono 16 kHz via ffmpeg.
    On décode nous-mêmes (au lieu de laisser Whisper appeler 'ffmpeg' dans le PATH),
    ce qui permet d'utiliser le ffmpeg fourni par pip — donc zéro installation système."""
    if not FFMPEG:
        raise FileNotFoundError("ffmpeg introuvable — réinstallez VoxLux.")
    cmd = [FFMPEG, '-nostdin', '-threads', '0', '-i', path,
           '-f', 's16le', '-ac', '1', '-acodec', 'pcm_s16le', '-ar', str(sr), '-']
    out = subprocess.run(cmd, capture_output=True, check=True, **_NO_WINDOW).stdout
    return np.frombuffer(out, np.int16).astype(np.float32) / 32768.0

# ─── Traitement des segments ──────────────────────────────────────────────────

def split_segment(seg, max_chars=80):
    """Découpe un segment trop long en sous-segments (récursif)."""
    text = seg['text'].strip()
    if len(text) <= max_chars:
        return [seg]
    mid = len(text) // 2
    split_at = -1
    for d in range(35):
        for i in [mid + d, mid - d]:
            if 0 < i < len(text) and text[i] in '.!?,; ':
                split_at = i + 1 if text[i] != ' ' else i
                break
        if split_at != -1:
            break
    if split_at == -1:
        split_at = mid
    t1 = text[:split_at].strip()
    t2 = text[split_at:].strip()
    if not t2:
        return [seg]
    ratio = split_at / len(text)
    mid_t = round(seg['start'] + (seg['end'] - seg['start']) * ratio, 2)
    return (split_segment({**seg, 'end': mid_t, 'text': t1}, max_chars) +
            split_segment({**seg, 'start': mid_t, 'text': t2}, max_chars))

# ─── Découpage intelligent par pauses ────────────────────────────────────────

def smart_split_words(words, min_pause=0.4, max_chars=60, min_duration=1.2):
    """Regroupe des mots en segments en coupant aux pauses et à la longueur max."""
    if not words:
        return []
    segments = []
    current = [words[0]]
    for word in words[1:]:
        pause = word['start'] - current[-1]['end']
        preview = ' '.join(w['word'] for w in current + [word]).strip()
        cut = (
            pause >= min_pause or
            len(preview) > max_chars or
            (pause >= 0.15 and current[-1]['word'].rstrip().endswith(('.', '!', '?', ';')))
        )
        if cut:
            text = ' '.join(w['word'] for w in current).strip()
            segments.append({'start': round(current[0]['start'], 2),
                             'end':   round(current[-1]['end'], 2),
                             'text':  text})
            current = [word]
        else:
            current.append(word)
    if current:
        text = ' '.join(w['word'] for w in current).strip()
        segments.append({'start': round(current[0]['start'], 2),
                         'end':   round(current[-1]['end'], 2),
                         'text':  text})
    # Durée minimale : évite les sous-titres qui disparaissent trop vite
    for i, seg in enumerate(segments):
        if seg['end'] - seg['start'] < min_duration:
            cap = segments[i + 1]['start'] if i + 1 < len(segments) else seg['start'] + min_duration + 1
            seg['end'] = round(min(seg['start'] + min_duration, cap), 2)
    for i, s in enumerate(segments):
        s['id'] = i
    return segments


def align_script(script_text, words):
    """Aligne les phrases d'un script sur les timestamps mot-à-mot de Whisper."""
    import re as _re
    # Découper le script en phrases
    raw = _re.split(r'(?<=[.!?…])\s+|\n+', script_text)
    sentences = [s.strip() for s in raw if s.strip() and len(s.strip()) > 2]
    if not words or not sentences:
        return []

    def norm(w):
        return _re.sub(r'[^\w]', '', w.lower())

    wtoks = [(norm(w['word']), w['start'], w['end']) for w in words if norm(w['word'])]
    n = len(wtoks)
    result = []
    last_pos = 0

    for sent in sentences:
        sw = [norm(w) for w in sent.split() if norm(w)]
        if not sw:
            continue
        span = len(sw)
        best_score = -1
        bi = last_pos
        bj = min(last_pos + span, n - 1)
        search_end = min(n, last_pos + span * 5 + 15)

        for si in range(last_pos, search_end):
            for ej in range(si, min(n, si + span * 2 + 8)):
                window = {wtoks[k][0] for k in range(si, ej + 1)}
                overlap = len(window & set(sw)) / max(len(sw), 1)
                lpen = abs((ej - si + 1) - span) / max(span, 1)
                score = overlap - 0.2 * lpen
                if score > best_score:
                    best_score = score
                    bi = si
                    bj = ej

        result.append({
            'id': len(result),
            'start': round(wtoks[bi][1], 2),
            'end':   round(wtoks[min(bj, n-1)][2], 2),
            'text':  sent,
        })
        last_pos = bj + 1

    return result

# ─── Formats d'export ─────────────────────────────────────────────────────────

def _srt_ts(s):
    h = int(s) // 3600; m = (int(s) % 3600) // 60; sec = s % 60
    return f"{h:02d}:{m:02d}:{sec:06.3f}".replace('.', ',')

def _vtt_ts(s):
    h = int(s) // 3600; m = (int(s) % 3600) // 60; sec = s % 60
    return f"{h:02d}:{m:02d}:{sec:06.3f}"

def make_srt(segs):
    return "\n".join(
        f"{i+1}\n{_srt_ts(s['start'])} --> {_srt_ts(s['end'])}\n{s['text']}\n"
        for i, s in enumerate(segs)
    )

def make_vtt(segs):
    parts = ["WEBVTT\n"]
    for s in segs:
        parts.append(f"{_vtt_ts(s['start'])} --> {_vtt_ts(s['end'])}\n{s['text']}\n")
    return "\n".join(parts)

def make_xlsx(segs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timecodes"
    hf = PatternFill("solid", start_color="1A1A2E")
    af = PatternFill("solid", start_color="F0F4FF")
    bd = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for col, title in enumerate(["Timecode", "Texte"], 1):
        c = ws.cell(1, col, title)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, seg in enumerate(segs):
        row = i + 2
        s = seg['start']
        m_total = int(s) // 60
        sec = int(s) % 60
        ms = round((s % 1) * 1000)
        tc = ws.cell(row, 1, f"{m_total:02d}:{sec:02d}:{ms:03d}")
        tx = ws.cell(row, 2, seg['text'])
        for c in [tc, tx]:
            c.font = Font(name="Calibri", size=10)
            c.border = bd
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 1:
                c.fill = af
        tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    path = tempfile.mktemp(suffix=".xlsx")
    wb.save(path)
    return path

# ─── Routes API ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return HTML, 200, {'Content-Type': 'text/html; charset=utf-8'}

@app.route('/api/upload', methods=['POST'])
def upload():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    fid = str(uuid.uuid4())
    ext = Path(f.filename).suffix.lower()
    path = os.path.join(UPLOAD_DIR, fid + ext)
    f.save(path)
    return jsonify({'file_id': fid, 'filename': f.filename, 'ext': ext})

MIME_TYPES = {
    '.mp4': 'video/mp4', '.mov': 'video/quicktime', '.avi': 'video/x-msvideo',
    '.mkv': 'video/x-matroska', '.webm': 'video/webm', '.m4v': 'video/mp4',
    '.mp3': 'audio/mpeg', '.wav': 'audio/wav', '.m4a': 'audio/mp4',
    '.ogg': 'audio/ogg', '.flac': 'audio/flac',
}

@app.route('/api/media/<fid>')
def serve_media(fid):
    path = next(
        (os.path.join(UPLOAD_DIR, n) for n in os.listdir(UPLOAD_DIR) if n.startswith(fid)),
        None
    )
    if not path:
        return 'Not found', 404

    file_size = os.path.getsize(path)
    ext = Path(path).suffix.lower()
    mime = MIME_TYPES.get(ext, 'application/octet-stream')
    range_header = request.headers.get('Range')

    if range_header:
        byte_start, byte_end = 0, file_size - 1
        m = re.search(r'bytes=(\d+)-(\d*)', range_header)
        if m:
            byte_start = int(m.group(1))
            if m.group(2):
                byte_end = int(m.group(2))
        length = byte_end - byte_start + 1

        def generate():
            with open(path, 'rb') as fh:
                fh.seek(byte_start)
                remaining = length
                while remaining > 0:
                    chunk = fh.read(min(65536, remaining))
                    if not chunk:
                        break
                    remaining -= len(chunk)
                    yield chunk

        return Response(generate(), 206, headers={
            'Content-Range': f'bytes {byte_start}-{byte_end}/{file_size}',
            'Accept-Ranges': 'bytes',
            'Content-Length': str(length),
            'Content-Type': mime,
        })

    return send_file(path, mimetype=mime, conditional=True)

def whisper_opts(langue, lexicon, word_timestamps=True):
    opts = {'task': 'transcribe', 'word_timestamps': word_timestamps}
    if langue and langue != 'auto':
        opts['language'] = langue
    # Évite qu'une fenêtre incertaine ne fasse "dérailler" les suivantes
    # (cause fréquente de passages entiers sautés au milieu d'une transcription).
    opts['condition_on_previous_text'] = False
    # Rend Whisper moins prompt à classer un passage "silence" (et donc à le
    # sauter entièrement) quand il n'est pas très confiant — voix faible,
    # accent, bruit de fond. Toujours probabiliste : n'élimine pas 100% des cas.
    opts['no_speech_threshold'] = 0.3   # défaut 0.6 — plus bas = moins de sauts pour "silence"
    opts['logprob_threshold']   = -1.5  # défaut -1.0 — accepte des passages moins confiants
    if lexicon:
        opts['initial_prompt'] = lexicon
    return opts

def find_upload_path(fid):
    return next(
        (os.path.join(UPLOAD_DIR, n) for n in os.listdir(UPLOAD_DIR) if n.startswith(fid)),
        None
    )

@app.route('/api/transcribe', methods=['POST'])
def transcribe():
    data = request.json
    fid = data.get('file_id')
    langue = data.get('langue', 'fr')
    path = find_upload_path(fid)
    if not path:
        return jsonify({'error': 'Fichier introuvable'}), 404
    try:
        audio = load_audio(path)
        lexicon = data.get('lexicon', '').strip()
        opts = whisper_opts(langue, lexicon)
        result = model.transcribe(audio, **opts, verbose=False)
        segs = []
        for i, s in enumerate(result.get('segments', [])):
            raw = {
                'id': i,
                'start': round(s['start'], 2),
                'end': round(s['end'], 2),
                'text': s['text'].strip(),
            }
            segs.extend(split_segment(raw))
        for i, s in enumerate(segs):
            s['id'] = i
        # Mots avec timestamps (pour le découpage intelligent)
        words = []
        for seg in result.get('segments', []):
            for w in seg.get('words', []):
                words.append({
                    'word':  w.get('word', ''),
                    'start': round(w.get('start', 0), 3),
                    'end':   round(w.get('end', 0), 3),
                })
        return jsonify({'segments': segs, 'words': words, 'detected_language': result.get('language', '')})
    except FileNotFoundError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/retranscribe_segment', methods=['POST'])
def retranscribe_segment():
    """Relance Whisper uniquement sur la plage [start, end] d'une ligne.
    Utile quand un passage a été sauté ou mal transcrit lors de la
    transcription complète — Whisper étant probabiliste, une nouvelle
    tentative isolée sur ce seul passage donne souvent un meilleur résultat."""
    data = request.json or {}
    fid = data.get('file_id')
    try:
        start = float(data.get('start', 0))
        end = float(data.get('end', 0))
    except (TypeError, ValueError):
        return jsonify({'error': 'Timecodes invalides'}), 400
    if end <= start:
        return jsonify({'error': 'Plage de temps invalide'}), 400
    path = find_upload_path(fid)
    if not path:
        return jsonify({'error': 'Fichier audio introuvable — relancez une transcription complète'}), 404
    try:
        audio = load_audio(path)
        sr = 16000
        pad = 0.4  # petit contexte avant/après pour aider Whisper à bien démarrer/finir
        i0 = max(0, int((start - pad) * sr))
        i1 = min(len(audio), int((end + pad) * sr))
        clip = audio[i0:i1]
        if len(clip) < int(sr * 0.2):
            return jsonify({'error': 'Plage trop courte pour être retranscrite'}), 400
        langue  = data.get('langue', 'fr')
        lexicon = (data.get('lexicon') or '').strip()
        opts = whisper_opts(langue, lexicon, word_timestamps=False)
        result = model.transcribe(clip, **opts, verbose=False)
        text = ' '.join(s['text'].strip() for s in result.get('segments', [])).strip()
        return jsonify({'text': text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/translate', methods=['POST'])
def translate():
    data = request.json
    segs = data.get('segments', [])
    lang = data.get('target_lang', 'fr')
    try:
        from deep_translator import GoogleTranslator
        tr = GoogleTranslator(source='auto', target=lang)
        result = []
        for s in segs:
            try:
                t = tr.translate(s['text'])
                result.append({**s, 'text': t or s['text']})
            except Exception:
                result.append(s)
        return jsonify({'segments': result})
    except ImportError:
        return jsonify({'error': "deep-translator non installé. Relancez l'installateur."}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/smart_split', methods=['POST'])
def smart_split_route():
    data = request.json
    words        = data.get('words', [])
    min_pause    = float(data.get('min_pause', 0.4))
    max_chars    = int(data.get('max_chars', 60))
    min_duration = float(data.get('min_duration', 1.2))
    if not words:
        return jsonify({'error': 'Aucun mot disponible — relancez la transcription'}), 400
    segs = smart_split_words(words, min_pause, max_chars, min_duration)
    return jsonify({'segments': segs})

@app.route('/api/align_script', methods=['POST'])
def align_script_route():
    data        = request.json
    script_text = data.get('script', '').strip()
    words_data  = data.get('words', [])
    if not script_text:
        return jsonify({'error': 'Script vide'}), 400
    if not words_data:
        return jsonify({'error': 'Aucun timestamp de mots — relancez la transcription'}), 400
    segs = align_script(script_text, words_data)
    return jsonify({'segments': segs})

@app.route('/api/export/srt', methods=['POST'])
def export_srt():
    segs = request.json.get('segments', [])
    return Response(make_srt(segs), mimetype='text/plain',
                    headers={'Content-Disposition': 'attachment; filename="subtitles.srt"'})

@app.route('/api/export/vtt', methods=['POST'])
def export_vtt():
    segs = request.json.get('segments', [])
    return Response(make_vtt(segs), mimetype='text/vtt',
                    headers={'Content-Disposition': 'attachment; filename="subtitles.vtt"'})

@app.route('/api/export/xlsx', methods=['POST'])
def export_xlsx():
    segs = request.json.get('segments', [])
    path = make_xlsx(segs)
    return send_file(path, as_attachment=True, download_name='transcription.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    """Arrête proprement le serveur VoxLux (bouton « Quitter » de l'interface)."""
    def _kill():
        import time
        time.sleep(0.3)
        os._exit(0)
    threading.Thread(target=_kill, daemon=True).start()
    return jsonify({'ok': True})

@app.route('/api/project/save', methods=['POST'])
def project_save():
    """Enregistre (ou met à jour) un projet en un vrai fichier sur disque.
    Appelé à la fois par le bouton « Enregistrer » et par l'autosave périodique."""
    data = request.json or {}
    project = data.get('project')
    if not isinstance(project, dict):
        return jsonify({'error': 'Projet invalide'}), 400
    name = safe_project_name(data.get('name'))
    old_filename = data.get('old_filename') or ''
    filename = name + PROJECT_SUFFIX
    path = PROJECTS_DIR / filename
    # Si ce nom appartient déjà à un AUTRE projet (pas celui qu'on met à jour),
    # on évite d'écraser en suffixant — comme le fait le Finder.
    if filename != old_filename and path.exists():
        i = 2
        while (PROJECTS_DIR / (name + f' ({i})' + PROJECT_SUFFIX)).exists():
            i += 1
        filename = name + f' ({i})' + PROJECT_SUFFIX
        path = PROJECTS_DIR / filename
    project['name'] = name
    project['savedAt'] = datetime.now().isoformat()
    try:
        path.write_text(json.dumps(project, ensure_ascii=False, indent=2), encoding='utf-8')
    except OSError as e:
        return jsonify({'error': str(e)}), 500
    # Renommage : on retire l'ancien fichier une fois le nouveau écrit avec succès
    if old_filename and old_filename != filename:
        old_path = PROJECTS_DIR / old_filename
        if old_path.exists():
            try: old_path.unlink()
            except OSError: pass
    return jsonify({'ok': True, 'filename': filename, 'name': name})

@app.route('/api/project/list')
def project_list():
    items = []
    for p in PROJECTS_DIR.glob('*' + PROJECT_SUFFIX):
        try:
            data = json.loads(p.read_text(encoding='utf-8'))
        except (OSError, ValueError):
            continue
        items.append({
            'filename': p.name,
            'name': data.get('name') or project_display_name(p),
            'savedAt': data.get('savedAt', ''),
            'segCount': len(data.get('segs') or []),
            'sourceFilename': data.get('sourceFilename', ''),
        })
    items.sort(key=lambda x: x['savedAt'], reverse=True)
    return jsonify({'projects': items[:30]})

@app.route('/api/project/load')
def project_load():
    filename = request.args.get('filename', '')
    if '/' in filename or '\\' in filename or '..' in filename or not filename.endswith(PROJECT_SUFFIX):
        return jsonify({'error': 'Nom de fichier invalide'}), 400
    path = PROJECTS_DIR / filename
    if not path.exists():
        return jsonify({'error': 'Projet introuvable'}), 404
    try:
        return jsonify(json.loads(path.read_text(encoding='utf-8')))
    except (OSError, ValueError) as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update/check')
def update_check():
    """Compare la version embarquée à la dernière release publiée sur GitHub.
    Échec réseau/API = silencieux (pas de connexion, pas grave, pas d'auto-update)."""
    asset_name = 'VoxLux-mac.dmg' if platform.system() == 'Darwin' else 'VoxLux-windows.zip'
    try:
        req = urllib.request.Request(
            f'https://api.github.com/repos/{GITHUB_REPO}/releases/latest',
            headers={'Accept': 'application/vnd.github+json', 'User-Agent': 'VoxLux-App'}
        )
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        latest = (data.get('tag_name') or '').lstrip('vV')
        asset_url = next((a.get('browser_download_url') for a in data.get('assets', [])
                           if a.get('name') == asset_name), None)
        return jsonify({
            'current': APP_VERSION,
            'latest': latest,
            'available': _version_tuple(latest) > _version_tuple(APP_VERSION),
            'asset_url': asset_url,
            'asset_name': asset_name,
            'notes_url': data.get('html_url', ''),
        })
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, ValueError, OSError) as e:
        return jsonify({'current': APP_VERSION, 'available': False, 'error': str(e)})

@app.route('/api/update/download', methods=['POST'])
def update_download():
    """Télécharge la mise à jour via curl (jamais via navigateur : un fichier
    téléchargé par le navigateur est marqué « quarantaine » par macOS et
    Gatekeeper en bloque l'ouverture — curl ne pose jamais cet attribut)."""
    data = request.json or {}
    url = data.get('url') or ''
    name = data.get('name') or 'VoxLux-update'
    version = data.get('version') or ''
    if not url.startswith('https://github.com/') and not url.startswith('https://objects.githubusercontent.com/'):
        return jsonify({'error': 'URL de mise à jour invalide'}), 400
    dest_dir = Path.home() / 'Downloads'
    dest_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(name).suffix or ''
    stem = Path(name).stem or 'VoxLux-update'
    dest = dest_dir / f"{stem}-v{version}{suffix}"
    try:
        subprocess.run(['curl', '-L', '-sS', '-o', str(dest), url], check=True, timeout=180)
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired, OSError) as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True, 'path': str(dest), 'filename': dest.name})

# ─── Interface HTML ───────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>VoxLux</title>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
      background: #07070f;
      color: white;
      height: 100vh;
      overflow: hidden;
      display: flex;
      flex-direction: column;
    }

    /* ── Gradient de fond animé + réactif souris ── */
    .bg {
      position: fixed; inset: 0; z-index: 0; pointer-events: none;
    }

    /* Grain */
    .noise {
      position: fixed; inset: 0; z-index: 1; pointer-events: none;
      opacity: 0.18;
      background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.65' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
      background-size: 220px 220px;
    }

    /* Vignette */
    .vignette {
      position: fixed; inset: 0; z-index: 2; pointer-events: none;
      background: radial-gradient(ellipse 80% 80% at 50% 50%,
        transparent 35%, rgba(4,4,10,0.72) 100%);
    }

    /* ── Tout le UI ── */
    .ui {
      position: relative; z-index: 10;
      display: flex; flex-direction: column; height: 100vh;
    }

    /* ══════════════════════════════
       HEADER commun
    ══════════════════════════════ */
    header {
      height: 58px; flex-shrink: 0;
      padding: 0 28px;
      display: flex; align-items: center;
      background: rgba(7,7,15,0.45);
      backdrop-filter: blur(24px);
      -webkit-backdrop-filter: blur(24px);
      border-bottom: 1px solid rgba(255,255,255,0.07);
    }

    .logo {
      font-size: 1.25em; font-weight: 800; letter-spacing: -0.05em;
      color: white; padding-right: 28px;
      border-right: 1px solid rgba(255,255,255,0.08);
      margin-right: 28px;
      display: flex; align-items: center; gap: 9px;
      flex-shrink: 0;
    }
    .logo-gem {
      width: 9px; height: 9px; border-radius: 2px;
      background: linear-gradient(135deg, #a78bfa, #ec4899);
      box-shadow: 0 0 12px rgba(167,139,250,0.7);
      transform: rotate(45deg); flex-shrink: 0;
    }

    .quit-btn {
      margin-left: auto; flex-shrink: 0;
      background: rgba(255,255,255,0.06);
      border: 1px solid rgba(255,255,255,0.12);
      color: rgba(255,255,255,0.55);
      font-weight: 700; font-size: 0.82em; letter-spacing: -0.02em;
      padding: 7px 15px; border-radius: 9px; cursor: pointer;
      transition: all 0.15s;
    }
    .quit-btn:hover {
      background: rgba(225,29,72,0.16); border-color: rgba(244,63,94,0.45);
      color: #fda4af;
    }

    .steps { display: flex; align-items: center; margin: 0 auto; }
    .step {
      display: flex; align-items: center; gap: 9px;
      padding: 0 20px; height: 58px;
      font-size: 0.88em; font-weight: 700; letter-spacing: -0.035em;
      color: rgba(255,255,255,0.28); cursor: default;
      position: relative; white-space: nowrap; transition: color 0.2s;
    }
    .step.done  { color: rgba(255,255,255,0.45); cursor: pointer; }
    .step.done:hover { color: rgba(255,255,255,0.65); }
    .step.active { color: rgba(255,255,255,0.92); }
    .step.active::after {
      content: ''; position: absolute; bottom: 0; left: 20px; right: 20px;
      height: 1.5px;
      background: linear-gradient(90deg, #a78bfa, #ec4899);
      border-radius: 2px 2px 0 0;
    }
    .stepnum {
      width: 22px; height: 22px; border-radius: 50%;
      background: rgba(255,255,255,0.07);
      display: flex; align-items: center; justify-content: center;
      font-size: 0.8em; font-weight: 800;
      color: rgba(255,255,255,0.3); flex-shrink: 0;
    }
    .step.active .stepnum {
      background: linear-gradient(135deg, #7c3aed, #db2777);
      color: white; box-shadow: 0 0 14px rgba(124,58,237,0.5);
    }
    .step.done .stepnum { background: rgba(167,139,250,0.2); color: #a78bfa; }
    .dash { color: rgba(255,255,255,0.1); padding: 0 2px; font-size: 0.8em; }

    /* ── Panels ── */
    .panel { display: none; flex: 1; flex-direction: column; overflow: hidden; min-height: 0; }
    .panel.active { display: flex; }

    /* ── Status bar ── */
    .spin {
      display: inline-block; width: 11px; height: 11px;
      border: 2px solid rgba(167,139,250,0.2); border-top-color: #a78bfa;
      border-radius: 50%; animation: rot 0.7s linear infinite; flex-shrink: 0;
    }
    @keyframes rot { to { transform: rotate(360deg); } }

    /* ── Boutons communs ── */
    .btn {
      padding: 8px 18px; border: none; border-radius: 10px;
      font-family: inherit; font-size: 0.88em; font-weight: 800;
      letter-spacing: -0.03em; cursor: pointer; white-space: nowrap;
      display: inline-flex; align-items: center; gap: 6px;
      transition: all 0.15s;
    }
    .btn:disabled { opacity: 0.25; cursor: not-allowed; transform: none !important; box-shadow: none !important; }
    .btn:not(:disabled):hover { transform: translateY(-1px); filter: brightness(1.12); }

    .btn-primary {
      background: linear-gradient(135deg, #7c3aed, #be185d);
      color: white; box-shadow: 0 0 20px rgba(124,58,237,0.35);
    }
    .btn-primary:not(:disabled):hover { box-shadow: 0 6px 28px rgba(124,58,237,0.5); }

    .btn-glass {
      background: rgba(255,255,255,0.07);
      border: 1px solid rgba(255,255,255,0.12);
      color: rgba(255,255,255,0.65);
    }
    .btn-glass:not(:disabled):hover { background: rgba(255,255,255,0.12); color: white; }

    .btn-green {
      background: linear-gradient(135deg, #065f46, #047857);
      color: rgba(255,255,255,0.9);
      border: 1px solid rgba(16,185,129,0.3);
    }
    .btn-green:not(:disabled):hover { box-shadow: 0 4px 16px rgba(5,150,105,0.4); }

    .btn-sm { padding: 6px 14px; font-size: 0.82em; border-radius: 8px; }

    /* ══════════════════════════════
       ÉTAPE 1 — UPLOAD
    ══════════════════════════════ */
    #p1 {
      align-items: center; justify-content: center;
    }

    .upload-card {
      width: 500px; max-width: 92vw;
      background: rgba(255,255,255,0.05);
      backdrop-filter: blur(32px);
      -webkit-backdrop-filter: blur(32px);
      border: 1px solid rgba(255,255,255,0.09);
      border-radius: 24px; padding: 38px 42px;
      box-shadow: inset 0 0 0 0.5px rgba(255,255,255,0.04),
                  0 32px 100px rgba(0,0,0,0.6),
                  0 4px 20px rgba(0,0,0,0.4);
      display: flex; flex-direction: column; gap: 0;
    }

    .card-title {
      font-size: 1.08em; font-weight: 800; letter-spacing: -0.045em;
      color: rgba(255,255,255,0.88); margin-bottom: 26px;
      display: flex; align-items: center; gap: 10px;
    }
    .title-bar {
      width: 3px; height: 16px; border-radius: 2px;
      background: linear-gradient(180deg, #a78bfa, #ec4899); flex-shrink: 0;
    }

    /* Bannière de mise à jour */
    .update-banner {
      display: flex; align-items: center; gap: 10px;
      background: linear-gradient(135deg, rgba(52,211,153,0.14), rgba(16,185,129,0.06));
      border: 1px solid rgba(52,211,153,0.3); border-radius: 12px;
      padding: 11px 14px; margin-bottom: 18px;
    }
    .update-banner-text {
      flex: 1; font-size: 0.82em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(167,243,208,0.95);
    }
    .update-banner-btn {
      flex-shrink: 0; background: rgba(52,211,153,0.85); border: none;
      color: #052e1e; font-size: 0.8em; font-weight: 800; letter-spacing: -0.02em;
      padding: 6px 14px; border-radius: 8px; cursor: pointer; transition: background 0.15s;
    }
    .update-banner-btn:hover { background: rgb(52,211,153); }
    .update-banner-btn:disabled { opacity: 0.6; cursor: default; }
    .update-banner-dismiss {
      flex-shrink: 0; background: none; border: none; cursor: pointer;
      color: rgba(167,243,208,0.5); font-size: 0.85em; padding: 2px 4px;
      transition: color 0.15s;
    }
    .update-banner-dismiss:hover { color: rgba(167,243,208,0.9); }

    /* Écran d'accueil */
    .landing-choice {
      display: flex; flex-direction: column; gap: 12px; margin-top: 22px;
    }
    .landing-btn {
      display: flex; align-items: center; gap: 16px; text-align: left;
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.09);
      border-radius: 14px; padding: 18px 20px; cursor: pointer;
      transition: background 0.15s, border-color 0.15s, transform 0.15s;
    }
    .landing-btn:hover { background: rgba(255,255,255,0.06); border-color: rgba(255,255,255,0.16); transform: translateY(-1px); }
    .landing-btn-primary {
      background: linear-gradient(135deg, rgba(124,58,237,0.16), rgba(190,24,93,0.1));
      border-color: rgba(167,139,250,0.3);
    }
    .landing-btn-primary:hover { border-color: rgba(167,139,250,0.55); }
    .landing-btn-emoji { font-size: 1.8em; flex-shrink: 0; }
    .landing-btn-text { display: flex; flex-direction: column; gap: 3px; }
    .landing-btn-title { font-size: 1em; font-weight: 800; letter-spacing: -0.03em; color: rgba(255,255,255,0.92); }
    .landing-btn-sub { font-size: 0.82em; font-weight: 600; color: rgba(255,255,255,0.4); }

    .back-link {
      background: none; border: none; cursor: pointer;
      font-size: 0.84em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.35); padding: 0 0 14px; display: block;
      transition: color 0.15s;
    }
    .back-link:hover { color: rgba(255,255,255,0.7); }

    .open-project-hint {
      font-size: 0.78em; font-weight: 600; color: rgba(255,255,255,0.32);
      margin: 10px 0 18px; text-align: center;
    }
    .action-row { margin-top: 22px; display: flex; justify-content: center; }
    .action-row .btn { min-width: 220px; }

    /* Dropzone */
    .dropzone {
      border: 1px dashed rgba(255,255,255,0.13);
      border-radius: 16px; padding: 34px 22px;
      text-align: center; cursor: pointer;
      background: rgba(255,255,255,0.02);
      transition: border-color 0.2s, background 0.2s;
      position: relative; overflow: hidden;
    }
    .dropzone::after {
      content: ''; position: absolute; inset: 0; pointer-events: none;
      background: radial-gradient(ellipse 80% 60% at 50% 0%,
        rgba(167,139,250,0.07) 0%, transparent 65%);
    }
    .dropzone:hover, .dropzone.over {
      border-color: rgba(167,139,250,0.4);
      background: rgba(167,139,250,0.03);
    }
    .dropzone.has-file { border-color: rgba(16,185,129,0.5); background: rgba(16,185,129,0.03); }

    .dz-icon { font-size: 2em; margin-bottom: 12px; opacity: 0.65; }
    .dz-main {
      font-size: 1em; font-weight: 800; letter-spacing: -0.04em;
      color: rgba(255,255,255,0.82); margin-bottom: 5px;
    }
    .dz-sub  { font-size: 0.82em; color: rgba(255,255,255,0.28); margin-bottom: 18px; }
    .dz-fname { font-size: 0.88em; font-weight: 700; color: #34d399; margin-top: 10px; display: none; }
    .dz-pick {
      display: inline-block;
      background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.14);
      border-radius: 9px; padding: 8px 20px;
      font-family: inherit; font-size: 0.86em; font-weight: 700; letter-spacing: -0.025em;
      color: rgba(255,255,255,0.65); cursor: pointer; transition: all 0.15s;
    }
    .dz-pick:hover { background: rgba(255,255,255,0.14); color: white; transform: translateY(-1px); }
    .dz-formats {
      margin-top: 14px; font-size: 0.73em; letter-spacing: 0.06em;
      color: rgba(255,255,255,0.18); font-weight: 600; text-transform: uppercase;
    }
    .dz-cancel-resume {
      display: block; margin: 14px auto 0; background: none; border: none; cursor: pointer;
      font-size: 0.78em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(244,63,94,0.6); padding: 4px 8px; transition: color 0.15s;
    }
    .dz-cancel-resume:hover { color: rgba(244,63,94,0.95); text-decoration: underline; }

    /* Ligne langue + transcription */
    .ctrl-row {
      display: flex; flex-direction: column; gap: 8px; margin-top: 22px;
    }
    .ctrl-label {
      font-size: 0.8em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.32);
    }
    .ctrl-inner { display: flex; align-items: center; gap: 10px; }
    .ctrl-select {
      flex: 1; background: rgba(255,255,255,0.06);
      border: 1px solid rgba(255,255,255,0.1); border-radius: 10px;
      padding: 9px 12px; font-family: inherit; font-size: 0.88em; font-weight: 600;
      color: rgba(255,255,255,0.72); cursor: pointer; appearance: none; outline: none;
      transition: border-color 0.15s;
    }
    .ctrl-select:focus { border-color: rgba(167,139,250,0.5); }

    /* Options avancées */
    .adv-toggle {
      width: 100%; background: none;
      border: 1px dashed rgba(255,255,255,0.07); border-radius: 9px;
      color: rgba(255,255,255,0.25); font-family: inherit; font-size: 0.82em;
      font-weight: 700; letter-spacing: -0.01em; padding: 8px 14px;
      cursor: pointer; text-align: left; transition: all 0.15s; margin-top: 20px;
    }
    .adv-toggle:hover, .adv-toggle.open {
      border-color: rgba(255,255,255,0.18); color: rgba(255,255,255,0.5);
    }
    .adv-panel { display: none; flex-direction: column; gap: 8px; margin-top: 12px; }
    .adv-panel.open { display: flex; }
    .adv-label {
      font-size: 0.84em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.45); display: flex; flex-direction: column; gap: 2px;
    }
    .adv-hint { font-weight: 500; font-size: 0.88em; color: rgba(255,255,255,0.25); }
    .adv-textarea {
      width: 100%; padding: 9px 12px;
      background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1);
      border-radius: 9px; font-family: inherit; font-size: 0.85em;
      color: rgba(255,255,255,0.7); resize: vertical; min-height: 60px; outline: none;
    }
    .adv-textarea:focus { border-color: rgba(167,139,250,0.4); }

    .card-note {
      margin-top: 20px; text-align: center;
      font-size: 0.76em; font-weight: 700; letter-spacing: -0.01em;
      color: rgba(255,255,255,0.18);
      display: flex; align-items: center; justify-content: center; gap: 7px;
    }
    .note-gem {
      width: 5px; height: 5px; border-radius: 1px;
      background: linear-gradient(135deg, #a78bfa, #ec4899);
      opacity: 0.5; transform: rotate(45deg); flex-shrink: 0;
    }

    .resume-row {
      margin-top: 14px; text-align: center;
      display: flex; align-items: center; justify-content: center; gap: 10px;
      flex-wrap: wrap;
    }
    .resume-link {
      background: none; border: none; cursor: pointer;
      font-size: 0.82em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(167,139,250,0.7); padding: 4px 2px;
      transition: color 0.15s;
    }
    .resume-link:hover { color: rgba(167,139,250,1); text-decoration: underline; }
    .resume-fname { font-size: 0.8em; font-weight: 700; color: #34d399; }

    .recent-projects { margin-top: 20px; }
    .recent-projects-title {
      font-size: 0.72em; font-weight: 800; letter-spacing: 0.06em; text-transform: uppercase;
      color: rgba(255,255,255,0.25); margin-bottom: 8px; text-align: center;
    }
    .recent-projects-list { display: flex; flex-direction: column; gap: 6px; }
    .recent-project-item {
      display: flex; align-items: center; gap: 10px;
      background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.07);
      border-radius: 10px; padding: 9px 14px; cursor: pointer;
      transition: background 0.15s, border-color 0.15s;
    }
    .recent-project-item:hover { background: rgba(167,139,250,0.08); border-color: rgba(167,139,250,0.25); }
    .recent-project-name { font-weight: 700; letter-spacing: -0.02em; color: rgba(255,255,255,0.85); flex: 1;
      overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .recent-project-meta { font-size: 0.78em; font-weight: 600; color: rgba(255,255,255,0.35); flex-shrink: 0; }

    /* ══════════════════════════════
       ÉTAPE 2 — TRANSCRIPTION
    ══════════════════════════════ */
    #p2 { flex-direction: column; }

    .p2-main { display: flex; flex: 1; overflow: hidden; min-height: 0; }

    /* Panneau vidéo */
    .vpanel {
      width: 48%; display: flex; flex-direction: column;
      background: #030307; flex-shrink: 0;
      border-right: 1px solid rgba(255,255,255,0.06);
    }
    .vwrap {
      flex: 1; position: relative;
      display: flex; align-items: center; justify-content: center; overflow: hidden;
    }
    video { max-width: 100%; max-height: 100%; display: block; }

    .sub-overlay {
      position: absolute; bottom: 90px; left: 50%; transform: translateX(-50%);
      color: #fff; font-size: 1.05em; font-weight: 700; letter-spacing: -0.02em;
      line-height: 1.45; text-align: center; max-width: 92%;
      background: rgba(0,0,0,0.6); padding: 5px 14px;
      border-radius: 6px; pointer-events: none; white-space: pre-line;
      text-shadow: 0 1px 6px rgba(0,0,0,0.9);
    }
    .tc-disp {
      position: absolute; top: 16px; left: 50%; transform: translateX(-50%);
      color: rgba(255,255,255,0.9); font-size: 1.5em;
      font-family: 'SF Mono', 'Fira Mono', monospace;
      background: rgba(0,0,0,0.55); padding: 5px 16px;
      border-radius: 8px; pointer-events: none; letter-spacing: 0.06em;
      font-weight: 600; white-space: nowrap;
    }
    .no-media {
      color: rgba(255,255,255,0.25); text-align: center; padding: 32px;
      font-size: 0.92em; font-weight: 600; letter-spacing: -0.02em; line-height: 1.7;
    }
    .nav-hint {
      position: absolute; bottom: 12px; left: 50%; transform: translateX(-50%);
      display: flex; gap: 14px; align-items: center;
      font-size: 0.82em; font-weight: 600; letter-spacing: -0.01em;
      color: rgba(255,255,255,0.38); white-space: nowrap; pointer-events: none;
    }
    .nav-hint-item { display: flex; align-items: center; gap: 5px; }
    .nav-key {
      display: inline-flex; align-items: center; justify-content: center;
      background: rgba(255,255,255,0.12); border: 1px solid rgba(255,255,255,0.2);
      border-radius: 5px; padding: 2px 8px; font-size: 0.92em;
      font-family: 'SF Mono', 'Fira Mono', monospace;
      color: rgba(255,255,255,0.55); line-height: 1.6;
    }
    .audio-wrap { padding: 28px; width: 100%; text-align: center; }
    .audio-wrap audio { width: 90%; margin-bottom: 14px; }
    .audio-sub {
      background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.08);
      border-radius: 10px; padding: 12px 18px; font-size: 0.95em; font-weight: 600;
      min-height: 48px; display: flex; align-items: center; justify-content: center;
      text-align: center; white-space: pre-line; color: rgba(255,255,255,0.7);
    }

    /* Panneau éditeur */
    .epanel {
      flex: 1; display: flex; flex-direction: column;
      background: rgba(7,7,15,0.7); min-width: 0;
    }

    .project-bar {
      padding: 8px 16px;
      background: rgba(124,58,237,0.06);
      border-bottom: 1px solid rgba(167,139,250,0.12);
      display: flex; align-items: center; gap: 8px; flex-shrink: 0;
    }
    .project-bar-icon { font-size: 0.9em; opacity: 0.6; flex-shrink: 0; }
    .project-name-input {
      background: none; border: none; outline: none;
      font-family: inherit; font-size: 0.9em; font-weight: 800; letter-spacing: -0.03em;
      color: rgba(255,255,255,0.85); flex: 1; min-width: 0;
      padding: 3px 6px; border-radius: 6px; transition: background 0.15s;
    }
    .project-name-input:hover  { background: rgba(255,255,255,0.05); }
    .project-name-input:focus  { background: rgba(167,139,250,0.1); }
    .project-save-state {
      font-size: 0.74em; font-weight: 700; color: rgba(255,255,255,0.28);
      flex-shrink: 0; white-space: nowrap;
    }
    .project-bar.unsaved .project-save-state { color: rgba(251,146,60,0.85); }
    .project-save-now-btn {
      flex-shrink: 0; background: rgba(251,146,60,0.16); border: 1px solid rgba(251,146,60,0.35);
      color: rgba(253,186,116,0.95); font-size: 0.76em; font-weight: 700; letter-spacing: -0.02em;
      padding: 4px 11px; border-radius: 7px; cursor: pointer; transition: background 0.15s;
    }
    .project-save-now-btn:hover { background: rgba(251,146,60,0.26); }

    /* ── Fenêtre de premier enregistrement ── */
    .save-prompt-overlay {
      position: fixed; inset: 0; z-index: 9998;
      display: none; align-items: center; justify-content: center;
      background: rgba(4,4,10,0.72); backdrop-filter: blur(6px); -webkit-backdrop-filter: blur(6px);
    }
    .save-prompt-overlay.open { display: flex; }
    .save-prompt-card {
      width: 92%; max-width: 420px;
      background: rgba(18,16,28,0.96); border: 1px solid rgba(167,139,250,0.22);
      border-radius: 18px; padding: 28px 26px;
      box-shadow: 0 30px 90px rgba(0,0,0,0.6);
    }
    .save-prompt-title {
      font-size: 1.2em; font-weight: 800; letter-spacing: -0.03em; color: white;
      margin-bottom: 18px;
    }
    .save-prompt-label {
      display: block; font-size: 0.78em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.4); margin-bottom: 7px;
    }
    .save-prompt-input {
      width: 100%; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.14);
      border-radius: 10px; padding: 10px 12px; font-family: inherit;
      font-size: 0.95em; font-weight: 700; color: white; outline: none;
      transition: border-color 0.15s;
    }
    .save-prompt-input:focus { border-color: rgba(167,139,250,0.55); }
    .save-prompt-location {
      font-size: 0.8em; font-weight: 600; color: rgba(255,255,255,0.38);
      margin: 12px 0 22px;
    }
    .save-prompt-location code {
      background: rgba(255,255,255,0.08); border-radius: 5px; padding: 1px 6px;
      color: rgba(167,139,250,0.85); font-size: 0.95em;
    }
    .save-prompt-actions { display: flex; align-items: center; gap: 16px; }
    .save-prompt-actions .btn { flex: 1; }
    .save-prompt-later {
      background: none; border: none; cursor: pointer;
      font-size: 0.84em; font-weight: 700; color: rgba(255,255,255,0.4);
      transition: color 0.15s;
    }
    .save-prompt-later:hover { color: rgba(255,255,255,0.7); }

    .etoolbar {
      padding: 10px 16px;
      background: rgba(255,255,255,0.03);
      border-bottom: 1px solid rgba(255,255,255,0.06);
      display: flex; align-items: center; gap: 8px; flex-shrink: 0; flex-wrap: wrap;
    }
    .etoolbar-label {
      font-size: 0.86em; font-weight: 800; letter-spacing: -0.035em;
      color: rgba(255,255,255,0.55);
    }
    .segcount {
      font-size: 0.78em; font-weight: 700; color: rgba(255,255,255,0.22);
      margin-left: auto;
    }

    /* Panneau découpage intelligent */
    .smart-bar {
      background: rgba(124,58,237,0.08); border-bottom: 1px solid rgba(167,139,250,0.15);
      padding: 12px 16px; font-size: 0.86em;
      display: none; flex-direction: column; gap: 10px; flex-shrink: 0;
    }
    .smart-bar.open { display: flex; }
    .smart-bar-row { display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }
    .smart-bar label {
      display: flex; align-items: center; gap: 8px; white-space: nowrap;
      color: rgba(255,255,255,0.55); font-weight: 700; letter-spacing: -0.02em;
    }
    .smart-bar input[type=range] { width: 100px; accent-color: #a78bfa; cursor: pointer; }
    .smart-bar .val { font-family: monospace; color: #a78bfa; min-width: 40px; }

    /* Toast de confirmation */
    .toast {
      position: fixed; bottom: 28px; left: 50%; transform: translateX(-50%) translateY(20px);
      background: rgba(15,15,25,0.92); backdrop-filter: blur(16px); -webkit-backdrop-filter: blur(16px);
      border: 1px solid rgba(167,139,250,0.3);
      color: white; font-weight: 700; letter-spacing: -0.02em; font-size: 0.88em;
      padding: 12px 22px; border-radius: 12px; z-index: 9999;
      box-shadow: 0 12px 40px rgba(0,0,0,0.5);
      opacity: 0; pointer-events: none; transition: opacity 0.25s, transform 0.25s;
    }
    .toast.show { opacity: 1; transform: translateX(-50%) translateY(0); }
    .smart-bar-note { font-size: 0.78em; color: rgba(255,255,255,0.25); font-weight: 600; }

    /* Panneau script */
    .script-bar {
      background: rgba(5,150,105,0.07); border-bottom: 1px solid rgba(16,185,129,0.15);
      padding: 12px 16px; font-size: 0.86em;
      display: none; flex-direction: column; gap: 10px; flex-shrink: 0;
    }
    .script-bar.open { display: flex; }
    .script-bar-row { display: flex; gap: 10px; align-items: flex-start; }
    .script-bar textarea {
      flex: 1; min-height: 80px; padding: 8px 12px;
      background: rgba(255,255,255,0.04); border: 1px solid rgba(16,185,129,0.2);
      border-radius: 8px; color: rgba(255,255,255,0.7);
      font-family: inherit; font-size: 0.87em; resize: vertical; line-height: 1.5; outline: none;
    }
    .script-bar textarea:focus { border-color: rgba(16,185,129,0.5); }
    .script-bar-actions { display: flex; flex-direction: column; gap: 8px; }
    .script-bar-note {
      font-size: 0.76em; color: rgba(255,255,255,0.25); font-weight: 600;
      line-height: 1.5; max-width: 300px;
    }

    /* Table des sous-titres */
    .etable-wrap { flex: 1; overflow-y: auto; }
    .empty-hint {
      height: 100%; display: flex; align-items: center; justify-content: center;
      color: rgba(255,255,255,0.2); font-size: 0.92em; font-weight: 600;
      letter-spacing: -0.02em; text-align: center; line-height: 1.8;
    }

    table { width: 100%; border-collapse: collapse; font-size: 0.88em; }
    thead th {
      background: rgba(255,255,255,0.04); color: rgba(255,255,255,0.4);
      padding: 8px 6px; text-align: left; position: sticky; top: 0; z-index: 1;
      font-weight: 700; font-size: 0.84em; letter-spacing: -0.02em;
      border-bottom: 1px solid rgba(255,255,255,0.06);
    }
    tbody tr {
      border-bottom: 1px solid rgba(255,255,255,0.04);
      cursor: pointer; transition: background 0.1s;
    }
    tbody tr:hover { background: rgba(255,255,255,0.04); }
    tbody tr.active { background: rgba(124,58,237,0.15) !important; }
    tbody tr.active td:first-child { border-left: 2px solid #a78bfa; }
    td { padding: 4px 4px; vertical-align: middle; }
    td:nth-child(4) { vertical-align: top; padding-top: 6px; }
    td.tnum { width: 28px; color: rgba(255,255,255,0.2); text-align: center; font-size: 0.78em; }
    td.ttime { width: 70px; }

    input.ti {
      width: 64px; padding: 4px 5px;
      background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.1);
      border-radius: 5px; font-size: 0.84em;
      font-family: 'SF Mono', 'Fira Mono', monospace;
      text-align: center; color: rgba(255,255,255,0.75); outline: none;
    }
    input.ti:focus { border-color: rgba(167,139,250,0.5); background: rgba(167,139,250,0.06); }

    textarea.tt {
      width: 100%; min-height: 36px; padding: 5px 7px;
      background: transparent; border: 1px solid transparent;
      border-radius: 5px; font-size: 0.88em; resize: none;
      font-family: inherit; line-height: 1.4; overflow: hidden;
      color: rgba(255,255,255,0.78); outline: none;
    }
    textarea.tt:focus {
      border-color: rgba(167,139,250,0.3);
      background: rgba(167,139,250,0.05);
    }
    textarea.tt.warn { border-color: rgba(245,158,11,0.4); background: rgba(245,158,11,0.04); }

    td.tdel { width: 52px; white-space: nowrap; }
    .delbtn, .rebtn {
      background: none; border: none; color: rgba(255,255,255,0.2);
      cursor: pointer; font-size: 0.9em; padding: 3px 5px;
      border-radius: 4px; line-height: 1; transition: all 0.15s;
    }
    .delbtn:hover { color: #f87171; background: rgba(248,113,113,0.1); }
    .rebtn:hover  { color: #a78bfa; background: rgba(167,139,250,0.12); }
    .rebtn:disabled { opacity: 0.4; cursor: default; }

    /* Footer étape 2 */
    .p2-footer {
      background: rgba(124,58,237,0.1);
      backdrop-filter: blur(12px);
      border-top: 1px solid rgba(167,139,250,0.22);
      padding: 11px 20px; display: flex; align-items: center;
      justify-content: center; gap: 14px; flex-wrap: wrap; flex-shrink: 0;
    }
    .p2-goto-bar {
      background: rgba(7,7,15,0.55);
      backdrop-filter: blur(12px);
      border-top: 1px solid rgba(255,255,255,0.05);
      padding: 16px 20px; display: flex; align-items: center;
      justify-content: center; flex-shrink: 0;
    }
    .fsec { display: flex; align-items: center; gap: 10px; }
    .flab { font-size: 0.84em; font-weight: 700; letter-spacing: -0.02em; color: rgba(255,255,255,0.35); }
    .fdiv { color: rgba(255,255,255,0.1); }
    .fmt-checks { display: flex; gap: 12px; align-items: center; }
    .fmt-checks label {
      display: flex; align-items: center; gap: 5px;
      font-size: 0.84em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.45); cursor: pointer; white-space: nowrap;
    }
    .fmt-checks input[type=checkbox] { accent-color: #a78bfa; cursor: pointer; }
    .push-right { margin-left: auto; }

    /* ══════════════════════════════
       ÉTAPE 3 — TRADUCTIONS
    ══════════════════════════════ */
    #p3 { flex-direction: column; }

    .p3-addbar {
      background: rgba(255,255,255,0.03);
      border-bottom: 1px solid rgba(255,255,255,0.06);
      padding: 12px 26px; display: flex; align-items: center;
      gap: 12px; flex-shrink: 0; flex-wrap: wrap;
    }
    .p3-addbar label {
      font-size: 0.86em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.35); white-space: nowrap;
    }
    .p3-addbar select {
      background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.1);
      border-radius: 9px; padding: 7px 12px;
      font-family: inherit; font-size: 0.88em; font-weight: 600;
      color: rgba(255,255,255,0.7); cursor: pointer; outline: none; appearance: none;
    }
    .p3-addbar-note { font-size: 0.76em; font-weight: 600; color: rgba(255,255,255,0.2); }

    .p3-cards-wrap {
      flex: 1; overflow-y: auto; padding: 20px 26px;
      display: flex; flex-direction: column; gap: 10px;
    }

    .p3-footer {
      background: rgba(167,139,250,0.06);
      backdrop-filter: blur(12px);
      border-top: 1px solid rgba(167,139,250,0.18);
      padding: 13px 26px; display: none; align-items: center;
      gap: 16px; flex-shrink: 0; flex-wrap: wrap;
    }
    .p3-footer.visible { display: flex; }
    .exp-checks { display: flex; gap: 12px; align-items: center; }
    .exp-checks label {
      display: flex; align-items: center; gap: 5px;
      font-size: 0.84em; font-weight: 700; letter-spacing: -0.02em;
      color: rgba(255,255,255,0.45); cursor: pointer;
    }
    .exp-checks input[type=checkbox] { accent-color: #a78bfa; cursor: pointer; }

    /* Cartes de langue */
    .lang-card {
      background: rgba(255,255,255,0.04);
      border: 1px solid rgba(255,255,255,0.08);
      border-radius: 14px; overflow: hidden;
      transition: border-color 0.2s;
    }
    .lang-card:hover { border-color: rgba(255,255,255,0.14); }

    .lang-card-header {
      padding: 13px 18px; display: flex; align-items: center; gap: 14px;
      cursor: pointer; user-select: none;
    }
    .lang-card-header:hover { background: rgba(255,255,255,0.02); }

    .lang-name {
      font-size: 0.95em; font-weight: 800; letter-spacing: -0.04em;
      color: rgba(255,255,255,0.8); min-width: 110px;
    }

    .lang-badge {
      font-size: 0.74em; font-weight: 800; padding: 3px 10px;
      border-radius: 20px; white-space: nowrap; letter-spacing: -0.01em;
    }
    .badge-wait    { background: rgba(167,139,250,0.12); color: rgba(167,139,250,0.8); }
    .badge-running { background: rgba(251,146,60,0.12); color: rgba(251,146,60,0.9); }
    .badge-done    { background: rgba(52,211,153,0.12); color: rgba(52,211,153,0.9); }
    .badge-error   { background: rgba(248,113,113,0.12); color: rgba(248,113,113,0.9); }

    .lang-exports { display: flex; gap: 7px; margin-left: auto; }
    .lang-expand { font-size: 0.8em; color: rgba(255,255,255,0.25); transition: transform 0.2s; }
    .lang-expand.open { transform: rotate(180deg); }
    .lang-remove {
      background: none; border: none; color: rgba(255,255,255,0.2);
      cursor: pointer; font-size: 1em; padding: 4px 6px;
      border-radius: 4px; transition: all 0.15s;
    }
    .lang-remove:hover { color: #f87171; background: rgba(248,113,113,0.1); }

    /* Prévisualisation avec comparaison */
    .lang-preview {
      display: none; border-top: 1px solid rgba(255,255,255,0.06);
      height: 260px; overflow-y: auto;
      background: rgba(0,0,0,0.15);
      /* scrollbar discrète */
      scrollbar-width: thin;
      scrollbar-color: rgba(167,139,250,0.25) transparent;
    }
    .lang-preview::-webkit-scrollbar { width: 4px; }
    .lang-preview::-webkit-scrollbar-track { background: transparent; }
    .lang-preview::-webkit-scrollbar-thumb { background: rgba(167,139,250,0.25); border-radius: 2px; }
    .lang-preview.open { display: block; }

    /* En-tête colonnes */
    .preview-header {
      display: grid; grid-template-columns: 72px 1fr 1fr;
      gap: 0; padding: 6px 18px;
      border-bottom: 1px solid rgba(255,255,255,0.08);
      font-size: 0.72em; font-weight: 800; letter-spacing: 0.06em;
      text-transform: uppercase;
    }
    .preview-header-tc { color: rgba(255,255,255,0.2); }
    .preview-header-orig { color: rgba(167,139,250,0.6); padding-left: 4px; }
    .preview-header-tr   { color: rgba(52,211,153,0.6); padding-left: 12px; border-left: 1px solid rgba(255,255,255,0.06); }

    .preview-seg {
      display: grid; grid-template-columns: 72px 1fr 1fr;
      gap: 0; padding: 6px 18px;
      border-bottom: 1px solid rgba(255,255,255,0.04);
      font-size: 0.84em; align-items: start;
    }
    .preview-seg:last-child { border-bottom: none; }
    .preview-tc {
      font-family: monospace; color: rgba(255,255,255,0.22);
      font-size: 0.84em; white-space: nowrap; padding-top: 1px;
    }
    .preview-orig {
      color: rgba(167,139,250,0.75); line-height: 1.45; font-weight: 500;
      padding: 0 8px 0 4px;
    }
    .preview-text {
      color: rgba(255,255,255,0.78); line-height: 1.45; font-weight: 500;
      padding: 2px 8px 2px 12px; border-left: 1px solid rgba(255,255,255,0.06);
      outline: none; border-radius: 0 4px 4px 0;
      transition: background 0.15s;
      cursor: text; white-space: pre-wrap; word-break: break-word;
    }
    .preview-text:focus {
      background: rgba(167,139,250,0.08);
      box-shadow: inset 2px 0 0 rgba(167,139,250,0.5);
    }
    .preview-text:hover:not(:focus) { background: rgba(255,255,255,0.03); }
    .preview-edit-hint {
      text-align: center; font-size: 0.72em; font-weight: 700;
      color: rgba(255,255,255,0.15); padding: 5px 0;
      letter-spacing: 0.03em; border-top: 1px solid rgba(255,255,255,0.04);
    }

    .lang-card-empty {
      text-align: center; color: rgba(255,255,255,0.2);
      font-size: 0.9em; font-weight: 700; letter-spacing: -0.02em;
      padding: 40px; border: 1px dashed rgba(255,255,255,0.08);
      border-radius: 14px;
    }
  </style>
</head>
<body>

<div class="bg" id="bg"></div>
<div class="noise"></div>
<div class="vignette"></div>

<div class="ui">

<!-- ── Header ── -->
<header>
  <div class="logo" onclick="goHome()" title="Retour à l'accueil" style="cursor:pointer;">
    <div class="logo-gem"></div>
    VoxLux
  </div>
  <div class="steps">
    <div class="step active" id="si1" onclick="stepClick(1)">
      <div class="stepnum"><span class="stepnum-n">1</span></div>
      <span class="steplabel">Upload</span>
    </div>
    <div class="dash">&#8212;</div>
    <div class="step" id="si2" onclick="stepClick(2)">
      <div class="stepnum"><span class="stepnum-n">2</span></div>
      <span class="steplabel" id="si2-label">Transcription</span>
    </div>
    <div class="dash">&#8212;</div>
    <div class="step" id="si3" onclick="stepClick(3)">
      <div class="stepnum"><span class="stepnum-n">3</span></div>
      <span class="steplabel">Traductions</span>
    </div>
  </div>
  <button class="quit-btn" id="quit-btn" title="Arr&#234;ter VoxLux" onclick="quitVoxLux()">&#9211; Quitter</button>
</header>

<!-- ══════════════════════════════
     ÉTAPE 1 — UPLOAD
══════════════════════════════ -->
<div class="panel active" id="p1">

  <!-- Écran d'accueil -->
  <div class="upload-card landing-card" id="landing-view">
    <div class="update-banner" id="update-banner" style="display:none;">
      <span class="update-banner-text" id="update-banner-text"></span>
      <button class="update-banner-btn" id="update-banner-btn" type="button">T&#233;l&#233;charger</button>
      <button class="update-banner-dismiss" id="update-banner-dismiss" type="button" title="Ignorer">&#10005;</button>
    </div>
    <div class="card-title">
      <div class="title-bar"></div>
      Bienvenue sur VoxLux
    </div>
    <div class="landing-choice">
      <button class="landing-btn landing-btn-primary" id="landing-new-btn" type="button">
        <span class="landing-btn-emoji">&#127909;</span>
        <span class="landing-btn-text">
          <span class="landing-btn-title">Nouveau projet</span>
          <span class="landing-btn-sub">Transcrire un nouveau fichier audio ou vid&#233;o</span>
        </span>
      </button>
      <button class="landing-btn" id="landing-open-btn" type="button">
        <span class="landing-btn-emoji">&#128193;</span>
        <span class="landing-btn-text">
          <span class="landing-btn-title">Ouvrir un projet existant</span>
          <span class="landing-btn-sub">Reprendre une transcription d&#233;j&#224; enregistr&#233;e</span>
        </span>
      </button>
    </div>
    <p class="card-note">
      <span class="note-gem"></span>
      Traitement 100&#37; local &mdash; aucun fichier envoy&#233; sur internet
    </p>
  </div>

  <!-- Espace de travail : nouveau projet OU reprise (structure partagée) -->
  <div class="upload-card" id="workspace-view" style="display:none;">
    <button class="back-link" id="back-to-landing-btn" type="button">&#8249; Retour</button>
    <div class="card-title">
      <div class="title-bar"></div>
      <span id="workspace-title-text">Nouveau fichier &#224; transcrire</span>
    </div>

    <!-- Visible uniquement en mode "Ouvrir un projet" -->
    <div id="open-project-extras" style="display:none;">
      <div class="recent-projects" id="recent-projects" style="display:none;">
        <div class="recent-projects-title">Projets r&#233;cents</div>
        <div class="recent-projects-list" id="recent-projects-list"></div>
      </div>
      <div class="resume-row">
        <input type="file" id="project-finput" accept=".json,.voxluxproj" style="display:none">
        <button class="resume-link" id="resume-project-btn" type="button">
          &#128194; Ou choisir un fichier projet (.json)
        </button>
        <span class="resume-fname" id="resume-fname"></span>
      </div>
      <p class="open-project-hint">Glissez ensuite la vid&#233;o d&#8217;origine ci-dessous, puis cliquez sur &#171;&nbsp;Reprendre cette transcription&nbsp;&#187;.</p>
    </div>

    <div class="dropzone" id="dropzone" onclick="document.getElementById('finput').click()">
      <input type="file" id="finput" accept=".mp3,.wav,.m4a,.ogg,.flac,.mp4,.mov,.avi,.mkv,.webm,.m4v" style="display:none">
      <div class="dz-icon" id="dz-icon"></div>
      <p class="dz-main" id="dz-main">Déposez votre fichier ici</p>
      <p class="dz-sub">ou parcourez vos fichiers</p>
      <button class="dz-pick" type="button" onclick="event.stopPropagation(); document.getElementById('finput').click()">Choisir un fichier</button>
      <p class="dz-formats">MP3 &middot; WAV &middot; M4A &middot; OGG &middot; FLAC &middot; MP4 &middot; MOV &middot; MKV &middot; WebM</p>
      <p class="dz-fname" id="dz-fname"></p>
      <button class="dz-cancel-resume" id="dz-cancel-resume" type="button"
        style="display:none;" onclick="event.stopPropagation(); cancelResume();">
        &#10005; Annuler la reprise &mdash; nouvelle transcription
      </button>
    </div>

    <!-- Visible uniquement en mode "Nouveau projet" -->
    <div id="new-project-extras">
      <div class="ctrl-row">
        <span class="ctrl-label">Langue de la vid&#233;o &#224; transcrire</span>
        <select class="ctrl-select" id="slang">
          <option value="fr">Fran&#231;ais</option>
          <option value="en">English</option>
          <option value="es">Espa&#241;ol</option>
          <option value="it">Italiano</option>
          <option value="de">Deutsch</option>
          <option value="pt">Portugu&#234;s</option>
          <option value="auto">D&#233;tection automatique</option>
        </select>
      </div>

      <button class="adv-toggle" id="adv-toggle" onclick="document.getElementById('adv-panel').classList.toggle('open');this.classList.toggle('open')">
        &#9881; Ajouter un lexique personnalis&#233;
      </button>
      <div class="adv-panel" id="adv-panel">
        <label class="adv-label" for="lexicon-input">
          Lexique personnalis&#233;
          <span class="adv-hint">Noms de marques, termes techniques — Whisper s'en servira pour mieux les reconna&#238;tre.</span>
        </label>
        <textarea id="lexicon-input" class="adv-textarea"
          placeholder="Un terme par ligne ou s&#233;par&#233;s par des virgules."></textarea>
      </div>
    </div>

    <div class="action-row">
      <button class="btn btn-primary" id="tbtn" disabled>&#9654; Transcrire</button>
    </div>

    <p class="card-note">
      <span class="note-gem"></span>
      Traitement 100&#37; local &mdash; aucun fichier envoy&#233; sur internet
    </p>
  </div>
</div>

<!-- ══════════════════════════════
     ÉTAPE 2 — TRANSCRIPTION
══════════════════════════════ -->
<div class="panel" id="p2">
  <div class="p2-main">

    <!-- Vidéo -->
    <div class="vpanel">
      <div class="vwrap" id="vwrap">
        <div class="no-media">Transcription en cours&#8230;</div>
      </div>
    </div>

    <!-- Éditeur -->
    <div class="epanel">
      <div class="project-bar" id="project-bar" style="display:none;">
        <span class="project-bar-icon">&#128193;</span>
        <input type="text" id="project-name-input" class="project-name-input" placeholder="Nom du projet" spellcheck="false">
        <span class="project-save-state" id="project-save-state"></span>
        <button class="project-save-now-btn" id="project-save-now-btn" type="button" style="display:none;">
          &#128190; Enregistrer
        </button>
      </div>
      <div class="etoolbar">
        <span class="etoolbar-label">Sous-titres</span>
        <button class="btn btn-glass btn-sm" id="addbtn" disabled>+ Ajouter</button>
        <button class="btn btn-sm" id="smart-toggle-btn"
          style="background:rgba(124,58,237,0.15);border:1px solid rgba(167,139,250,0.25);color:rgba(167,139,250,0.85);" disabled>
          &#9889; D&#233;coupage intelligent
        </button>
        <button class="btn btn-sm" id="script-toggle-btn"
          style="background:rgba(5,150,105,0.12);border:1px solid rgba(16,185,129,0.2);color:rgba(52,211,153,0.8);" disabled>
          &#128196; Aligner un script
        </button>
        <span class="segcount" id="segcount"></span>
      </div>

      <!-- Panneau découpage intelligent -->
      <div class="smart-bar" id="smart-bar">
        <div class="smart-bar-row">
          <label>
            Pause min :
            <input type="range" id="sr-pause" min="0.1" max="1.5" step="0.05" value="0.4">
            <span class="val" id="sr-pause-val">0.40 s</span>
          </label>
          <label>
            Longueur max :
            <input type="range" id="sr-chars" min="20" max="120" step="5" value="60">
            <span class="val" id="sr-chars-val">60 car.</span>
          </label>
          <button class="btn btn-sm" id="smart-apply-btn"
            style="background:linear-gradient(135deg,#7c3aed,#be185d);color:white;">
            Appliquer
          </button>
        </div>
        <div class="smart-bar-row">
          <label>
            Dur&#233;e min :
            <input type="range" id="sr-dur" min="0.5" max="4" step="0.1" value="1.2">
            <span class="val" id="sr-dur-val">1.2 s</span>
          </label>
        </div>
        <p class="smart-bar-note">Coupe aux pauses entre mots (&ge; pause min) et &#224; la longueur max.</p>
      </div>

      <!-- Panneau script -->
      <div class="script-bar" id="script-bar">
        <div class="script-bar-row">
          <textarea id="script-input" placeholder="Collez le script ici &mdash; une phrase par ligne ou s&#233;par&#233;es par . ! ?"></textarea>
          <div class="script-bar-actions">
            <button class="btn btn-sm" id="align-btn"
              style="background:linear-gradient(135deg,#065f46,#047857);color:white;">
              Aligner &#8594;
            </button>
            <p class="script-bar-note">
              L'app associe chaque phrase du script au timecode correspondant.<br>
              Les timecodes Whisper sont conserv&#233;s, le texte du script les remplace.
            </p>
          </div>
        </div>
      </div>

      <div class="etable-wrap">
        <div class="empty-hint" id="ehint">
          Transcription en cours&#8230;
        </div>
        <table id="stbl" style="display:none">
          <thead>
            <tr>
              <th class="tnum">#</th>
              <th style="width:70px">D&#233;but</th>
              <th style="width:70px">Fin</th>
              <th>Texte</th>
              <th style="width:28px"></th>
            </tr>
          </thead>
          <tbody id="stbody"></tbody>
        </table>
      </div>
    </div><!-- /epanel -->

  </div><!-- /p2-main -->

  <div class="p2-footer">
    <span class="flab">Exporter la transcription :</span>
    <div class="fmt-checks">
      <label><input type="checkbox" id="m-xlsx" checked> Excel</label>
      <label><input type="checkbox" id="m-srt"  checked> SRT</label>
      <label><input type="checkbox" id="m-vtt"  checked> WebVTT</label>
    </div>
    <button class="btn btn-primary btn-sm" id="master-export-btn" disabled>
      &#8595; T&#233;l&#233;charger transcription-<span id="export-lang-label">FR</span>
    </button>
  </div>
  <div class="p2-goto-bar">
    <button class="btn btn-sm" id="goto3-btn" disabled
      style="background:rgba(167,139,250,0.12);border:1px solid rgba(167,139,250,0.3);color:rgba(167,139,250,0.85);">
      Passer &#224; la traduction &#8594;
    </button>
  </div>
</div>

<!-- ══════════════════════════════
     ÉTAPE 3 — TRADUCTIONS
══════════════════════════════ -->
<div class="panel" id="p3">

  <div class="p3-addbar">
    <label>Ajouter une langue :</label>
    <select id="tlang-pick">
      <option value="fr">Fran&#231;ais</option>
      <option value="en">English</option>
      <option value="es">Espa&#241;ol</option>
      <option value="it">Italiano</option>
      <option value="de">Deutsch</option>
      <option value="pt">Portugu&#234;s</option>
      <option value="nl">Nederlands</option>
      <option value="ja">&#26085;&#26412;&#35486;</option>
      <option value="zh-CN">&#20013;&#25991;</option>
      <option value="ar">&#1575;&#1604;&#1593;&#1585;&#1576;&#1610;&#1577;</option>
      <option value="ru">&#1056;&#1091;&#1089;&#1089;&#1082;&#1080;&#1081;</option>
    </select>
    <button class="btn btn-glass btn-sm" id="add-lang-btn">+ Ajouter</button>
    <div style="width:1px;height:22px;background:rgba(255,255,255,0.08);margin:0 4px;"></div>
    <button class="btn btn-primary btn-sm" id="translate-all-btn" disabled>
      &#127760; Tout traduire
    </button>
    <span class="p3-addbar-note" style="margin-left:4px;">Cliquez sur une carte termin&#233;e pour comparer avec la transcription.</span>
  </div>

  <div class="p3-cards-wrap" id="lang-cards">
    <div class="lang-card-empty" id="lang-empty">
      Ajoutez des langues cibles ci-dessus pour commencer.
    </div>
  </div>

  <div class="p3-footer" id="p3-footer">
    <span style="font-size:0.82em;font-weight:700;letter-spacing:-0.02em;color:rgba(255,255,255,0.35);">Tout exporter :</span>
    <div class="exp-checks">
      <label><input type="checkbox" id="e-xlsx" checked> Excel</label>
      <label><input type="checkbox" id="e-srt"  checked> SRT</label>
      <label><input type="checkbox" id="e-vtt"  checked> WebVTT</label>
    </div>
    <button class="btn btn-green push-right" id="export-all-btn" disabled>
      &#8595; Tout exporter (termin&#233;es)
    </button>
  </div>

</div>

<!-- ── Status bar ── -->

</div><!-- /ui -->

<!-- ── Premier enregistrement du projet ── -->
<div class="save-prompt-overlay" id="save-prompt-overlay">
  <div class="save-prompt-card">
    <div class="save-prompt-title">&#9989; Transcription termin&#233;e</div>
    <label class="save-prompt-label" for="save-prompt-name">Nom du projet</label>
    <input type="text" id="save-prompt-name" class="save-prompt-input" spellcheck="false">
    <p class="save-prompt-location">Sera enregistr&#233; dans <code>Documents/VoxLux&nbsp;Projets/</code></p>
    <div class="save-prompt-actions">
      <button class="btn btn-primary" id="save-prompt-confirm-btn">&#128190; Enregistrer le projet</button>
      <button class="save-prompt-later" id="save-prompt-later-btn" type="button">Plus tard</button>
    </div>
  </div>
</div>

<script>
// ── État global ───────────────────────────────────────────────────────────
let fileId      = null;
let segs        = [];
let words       = [];
let media       = null;
let trs         = {};
let detectedLang = '';
let currentLangue  = 'fr';  // langue utilisée à la transcription — réutilisée pour "Retranscrire ce passage"
let currentLexicon = '';

const LANG_NAMES = {
  'fr':'Français','en':'English','es':'Español','it':'Italiano',
  'de':'Deutsch','pt':'Português','nl':'Nederlands',
  'ja':'日本語','zh-CN':'中文',
  'ar':'العربية','ru':'Русский'
};

// ── Gradient de fond réactif souris ──────────────────────────────────────
const bgEl = document.getElementById('bg');
let tX = 72, tY = 18, cX = 72, cY = 18;

function lerp(a, b, t) { return a + (b - a) * t; }

document.addEventListener('mousemove', e => {
  tX = (e.clientX / window.innerWidth)  * 100;
  tY = (e.clientY / window.innerHeight) * 100;
});

function tickBg() {
  cX = lerp(cX, tX, 0.04);
  cY = lerp(cY, tY, 0.04);
  const x2 = 100 - cX * 0.6, y2 = 100 - cY * 0.5;
  bgEl.style.background = [
    'radial-gradient(ellipse 140% 110% at ' + cX.toFixed(1) + '% ' + cY.toFixed(1) + '%, rgba(109,40,217,0.55) 0%, rgba(190,24,93,0.28) 42%, transparent 72%)',
    'radial-gradient(ellipse 130% 100% at ' + x2.toFixed(1) + '% ' + y2.toFixed(1) + '%, rgba(29,78,216,0.48) 0%, rgba(6,182,212,0.18) 45%, transparent 72%)',
    'radial-gradient(ellipse 100% 80% at ' + (100 - cX * 0.3).toFixed(1) + '% ' + (cY * 0.8 + 30).toFixed(1) + '%, rgba(124,58,237,0.20) 0%, transparent 65%)',
    'linear-gradient(160deg, #0d0818 0%, #07070f 50%, #080d18 100%)'
  ].join(',');
  requestAnimationFrame(tickBg);
}
tickBg();

// ── Navigation étapes ─────────────────────────────────────────────────────
function goTo(n) {
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.getElementById('p' + n).classList.add('active');
  [1,2,3].forEach(i => {
    const el = document.getElementById('si' + i);
    el.classList.remove('active','done');
    if (i < n)  el.classList.add('done');
    if (i === n) el.classList.add('active');
  });
}

function stepClick(n) {
  const el = document.getElementById('si' + n);
  if (!el.classList.contains('done') && !el.classList.contains('active')) return;
  if (n === 1 && segs.length > 0) {
    if (!confirm("Retourner à l'Upload effacera la transcription en cours. Continuer ?")) return;
  }
  goTo(n);
}

// Clic sur le logo « VoxLux » : retour direct à l'écran d'accueil (Nouveau
// projet / Ouvrir un projet), pas juste à l'étape 1.
function goHome() {
  if (segs.length > 0) {
    if (!confirm("Retourner à l'accueil effacera la transcription en cours. Continuer ?")) return;
  }
  goTo(1);
  showLanding();
}

// ── Quitter VoxLux ─────────────────────────────────────────────────────────
function showQuitOverlay(title, sub) {
  const o = document.createElement('div');
  o.style.cssText = 'position:fixed;inset:0;z-index:99999;display:flex;flex-direction:column;'
    + 'align-items:center;justify-content:center;gap:14px;background:rgba(7,7,15,0.94);'
    + 'backdrop-filter:blur(10px);-webkit-backdrop-filter:blur(10px);font-family:inherit;text-align:center;padding:0 30px;';
  o.innerHTML = '<div style="font-size:1.7em;font-weight:800;letter-spacing:-0.04em;color:#fff;">' + title + '</div>'
    + '<div style="font-weight:600;color:rgba(255,255,255,0.5);">' + sub + '</div>';
  document.body.appendChild(o);
}

async function quitVoxLux() {
  if (!confirm('Arrêter VoxLux ? Le travail non exporté sera perdu.')) return;
  try { await fetch('/api/shutdown', { method: 'POST' }); } catch (e) {}
  showQuitOverlay('VoxLux est arrêté', 'Vous pouvez fermer cet onglet. Relancez VoxLux depuis son icône.');
}

// ── Status ────────────────────────────────────────────────────────────────
function status(msg, loading) {
  // sbar supprimée — statut silencieux (erreurs restent en console)
  if (loading) console.log('⏳', msg); else console.log(msg);
}

// ── Toast de confirmation ────────────────────────────────────────────────
let toastTimer = null;
function toast(msg, duration) {
  let el = document.getElementById('toast');
  if (!el) {
    el = document.createElement('div');
    el.id = 'toast'; el.className = 'toast';
    document.body.appendChild(el);
  }
  el.textContent = msg;
  clearTimeout(toastTimer);
  requestAnimationFrame(() => el.classList.add('show'));
  toastTimer = setTimeout(() => el.classList.remove('show'), duration || 2600);
}

// ── Utilitaires temps ─────────────────────────────────────────────────────
function parseT(str) {
  str = str.trim();
  const p = str.split(':');
  if (p.length === 1) return parseFloat(p[0]);
  if (p.length === 2) return parseFloat(p[0]) * 60 + parseFloat(p[1]);
  return parseFloat(p[0]) * 3600 + parseFloat(p[1]) * 60 + parseFloat(p[2]);
}
function fmtT(s) {
  const m = Math.floor(s / 60);
  const sec = (s % 60).toFixed(1).padStart(4,'0');
  return String(m).padStart(2,'0') + ':' + sec;
}
function fmtTC(s) {
  const m  = Math.floor(s / 60);
  const sc = Math.floor(s % 60);
  const ms = Math.round((s % 1) * 1000);
  return String(m).padStart(2,'0') + ':' + String(sc).padStart(2,'0') + '.' + String(ms).padStart(3,'0');
}
function r1(n) { return Math.round(n * 10) / 10; }
function esc(s) {
  return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

// ── Upload & drag-drop ────────────────────────────────────────────────────
let selectedFile = null;
const dz = document.getElementById('dropzone');

document.getElementById('finput').addEventListener('change', e => {
  if (e.target.files[0]) handleFile(e.target.files[0]);
});
dz.addEventListener('dragenter', e => { e.preventDefault(); e.stopPropagation(); dz.classList.add('over'); });
dz.addEventListener('dragover',  e => { e.preventDefault(); e.stopPropagation(); dz.classList.add('over'); });
dz.addEventListener('dragleave', e => { e.stopPropagation(); dz.classList.remove('over'); });
dz.addEventListener('drop', e => {
  e.preventDefault(); e.stopPropagation(); dz.classList.remove('over');
  const f = e.dataTransfer.files[0]; if (f) handleFile(f);
});

function handleFile(file) {
  selectedFile = file;
  document.getElementById('dz-fname').textContent = '✅ ' + file.name;
  document.getElementById('dz-fname').style.display = 'block';
  document.getElementById('dz-icon').textContent = '';
  document.getElementById('dz-main').textContent = file.name;
  dz.classList.add('has-file');
  document.getElementById('tbtn').disabled = false;
  status('Fichier sélectionné : ' + file.name);
}

// ── Projet : nom, fichier réel sur disque, reprise ──────────────────────────
let resumeMode        = false;
let projectName       = '';
let projectFilename   = '';   // nom de fichier confirmé par le serveur (gère le renommage proprement)
let projectSavedOnce  = false; // un vrai fichier existe déjà sur disque pour ce projet

function setProjectName(name) {
  projectName = name || 'Projet sans titre';
  document.getElementById('project-name-input').value = projectName;
  document.getElementById('project-bar').style.display = 'flex';
}

// État visuel de la barre projet : « ✅ Enregistré à HH:MM » (vert) ou
// « ⚠️ Non enregistré » avec un bouton d'enregistrement immédiat (orange).
function markProjectSaved(whenISO) {
  const d = whenISO ? new Date(whenISO) : new Date();
  document.getElementById('project-bar').classList.remove('unsaved');
  document.getElementById('project-save-state').textContent =
    '✅ Enregistré ' + d.toLocaleTimeString('fr-FR', { hour: '2-digit', minute: '2-digit' });
  document.getElementById('project-save-now-btn').style.display = 'none';
}
function markProjectUnsaved() {
  document.getElementById('project-bar').classList.add('unsaved');
  document.getElementById('project-save-state').textContent = '⚠️ Non enregistré';
  document.getElementById('project-save-now-btn').style.display = '';
}

// ── Fenêtre de premier enregistrement (affichée à la fin d'une transcription) ──
function openSavePrompt(defaultName) {
  document.getElementById('save-prompt-name').value = defaultName || projectName || 'Projet sans titre';
  document.getElementById('save-prompt-overlay').classList.add('open');
  setTimeout(() => document.getElementById('save-prompt-name').select(), 60);
}
function closeSavePrompt() {
  document.getElementById('save-prompt-overlay').classList.remove('open');
}
document.getElementById('save-prompt-confirm-btn').addEventListener('click', async () => {
  const name = document.getElementById('save-prompt-name').value.trim() || 'Projet sans titre';
  setProjectName(name);
  closeSavePrompt();
  await saveProjectToServer();
});
document.getElementById('save-prompt-later-btn').addEventListener('click', () => {
  closeSavePrompt();
  markProjectUnsaved();
});
document.getElementById('save-prompt-name').addEventListener('keydown', e => {
  if (e.key === 'Enter') { e.preventDefault(); document.getElementById('save-prompt-confirm-btn').click(); }
});
document.getElementById('project-save-now-btn').addEventListener('click', () => {
  const v = document.getElementById('project-name-input').value.trim();
  if (v) projectName = v;
  saveProjectToServer();
});

function applyProjectData(data, msg) {
  segs = data.segs || [];
  words = data.words || [];
  trs = data.trs || {};
  detectedLang = data.detectedLang || '';
  window.__voxluxSourceName = data.sourceFilename || '';
  projectFilename = data.__filename || '';
  projectSavedOnce = true;   // un fichier existe déjà (ce chargement en est la preuve)
  resumeMode = true;

  setProjectName(data.name || (data.sourceFilename || 'Projet').replace(/\.[^.]+$/, ''));
  markProjectSaved(data.savedAt);
  document.getElementById('resume-fname').textContent = '✅ ' + segs.length + ' segments chargés';
  document.getElementById('tbtn').textContent = '▶ Reprendre cette transcription';
  document.getElementById('dz-main').textContent = 'Glissez la vidéo d’origine (' +
    (data.sourceFilename || 'même fichier que la transcription') + ')';
  document.getElementById('dz-cancel-resume').style.display = 'block';
  toast(msg || ('📂 Projet chargé — glissez la vidéo d’origine pour continuer'));
}

// Sortir du mode reprise pour repartir sur une transcription neuve
// Remise à zéro complète : aucune session (reprise ou transcription en cours)
// ne doit jamais survivre à un aller-retour vers l'écran d'accueil.
function resetProjectState() {
  resumeMode = false;
  segs = []; words = []; trs = {};
  selectedFile = null;
  document.getElementById('finput').value = '';
  dz.classList.remove('has-file');
  document.getElementById('dz-icon').textContent = '';
  document.getElementById('dz-main').textContent = 'Déposez votre fichier ici';
  document.getElementById('dz-fname').style.display = 'none';
  document.getElementById('dz-fname').textContent = '';
  document.getElementById('dz-cancel-resume').style.display = 'none';
  document.getElementById('resume-fname').textContent = '';
  document.getElementById('tbtn').textContent = '▶ Transcrire';
  document.getElementById('tbtn').disabled = true;
  document.getElementById('project-bar').style.display = 'none';
  document.getElementById('project-bar').classList.remove('unsaved');
  projectName = ''; projectFilename = ''; projectSavedOnce = false;
  clearAutosave();
}

function cancelResume() {
  resetProjectState();
  toast('Reprise annulée — prêt pour une nouvelle transcription');
}

document.getElementById('resume-project-btn').addEventListener('click', () => {
  document.getElementById('project-finput').click();
});
document.getElementById('project-finput').addEventListener('change', e => {
  const f = e.target.files[0]; if (!f) return;
  const reader = new FileReader();
  reader.onload = () => {
    let data;
    try { data = JSON.parse(reader.result); } catch (err) { toast('❌ Fichier projet invalide'); return; }
    if (!data.segs) { toast('❌ Fichier projet invalide'); return; }
    applyProjectData(data);
  };
  reader.readAsText(f);
  e.target.value = '';
});

// Renommer le projet (déclenche une sauvegarde immédiate sous le nouveau nom)
document.getElementById('project-name-input').addEventListener('blur', () => {
  const v = document.getElementById('project-name-input').value.trim();
  if (!v || v === projectName) { document.getElementById('project-name-input').value = projectName; return; }
  projectName = v;
  saveProjectToServer();
  toast('✏️ Projet renommé : ' + projectName);
});
document.getElementById('project-name-input').addEventListener('keydown', e => {
  if (e.key === 'Enter') { e.preventDefault(); document.getElementById('project-name-input').blur(); }
});

// ── Projets récents (étape 1) ───────────────────────────────────────────────
async function loadRecentProjects() {
  let data;
  try {
    const res = await fetch('/api/project/list');
    data = await res.json();
  } catch (e) { return; }
  const list = (data && data.projects) || [];
  const box = document.getElementById('recent-projects');
  const ul  = document.getElementById('recent-projects-list');
  if (!list.length) { box.style.display = 'none'; return; }
  ul.innerHTML = '';
  list.forEach(p => {
    const when = p.savedAt ? new Date(p.savedAt).toLocaleString('fr-FR', { day:'2-digit', month:'2-digit', hour:'2-digit', minute:'2-digit' }) : '';
    const item = document.createElement('div');
    item.className = 'recent-project-item';
    item.innerHTML =
      '<span class="recent-project-name">' + esc(p.name) + '</span>' +
      '<span class="recent-project-meta">' + p.segCount + ' seg. · ' + when + '</span>';
    item.addEventListener('click', async () => {
      let full;
      try {
        const res = await fetch('/api/project/load?filename=' + encodeURIComponent(p.filename));
        full = await res.json();
      } catch (e) { toast('❌ Impossible de charger ce projet'); return; }
      if (full.error) { toast('❌ ' + full.error); return; }
      full.__filename = p.filename;
      applyProjectData(full, '📂 « ' + p.name + ' » chargé — glissez la vidéo d’origine pour continuer');
    });
    ul.appendChild(item);
  });
  box.style.display = '';
}

// ── Écran d'accueil : Nouveau projet / Ouvrir un projet existant ───────────
function showLanding() {
  document.getElementById('landing-view').style.display = '';
  document.getElementById('workspace-view').style.display = 'none';
}
function showWorkspace(mode) {
  const isOpen = mode === 'open';
  document.getElementById('landing-view').style.display = 'none';
  document.getElementById('workspace-view').style.display = 'block';
  document.getElementById('open-project-extras').style.display = isOpen ? 'block' : 'none';
  document.getElementById('new-project-extras').style.display = isOpen ? 'none' : 'block';
  document.getElementById('workspace-title-text').textContent =
    isOpen ? 'Ouvrir un projet existant' : 'Nouveau fichier à transcrire';
  if (isOpen) loadRecentProjects();
}
document.getElementById('landing-new-btn').addEventListener('click', () => { resetProjectState(); showWorkspace('new'); });
document.getElementById('landing-open-btn').addEventListener('click', () => { resetProjectState(); showWorkspace('open'); });
document.getElementById('back-to-landing-btn').addEventListener('click', showLanding);

// ── Mise à jour automatique ─────────────────────────────────────────────────
// Vérification silencieuse au lancement : n'affiche une bannière que si une
// version plus récente existe réellement. Aucune interruption sinon.
let pendingUpdate = null;
async function checkForUpdate() {
  let data;
  try {
    const res = await fetch('/api/update/check');
    data = await res.json();
  } catch (e) { return; }
  if (!data || !data.available || !data.asset_url) return;
  pendingUpdate = data;
  document.getElementById('update-banner-text').textContent =
    '🔔 Mise à jour disponible — version ' + data.latest + ' (vous avez la ' + data.current + ')';
  document.getElementById('update-banner').style.display = 'flex';
}
document.getElementById('update-banner-dismiss').addEventListener('click', () => {
  document.getElementById('update-banner').style.display = 'none';
});
document.getElementById('update-banner-btn').addEventListener('click', async () => {
  if (!pendingUpdate) return;
  const btn = document.getElementById('update-banner-btn');
  btn.disabled = true; btn.textContent = 'Téléchargement…';
  try {
    const res = await fetch('/api/update/download', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ url: pendingUpdate.asset_url, name: pendingUpdate.asset_name, version: pendingUpdate.latest })
    });
    const result = await res.json();
    if (result.error) { toast('❌ ' + result.error); btn.disabled = false; btn.textContent = 'Télécharger'; return; }
    try { await fetch('/api/shutdown', { method: 'POST' }); } catch (e) {}
    showQuitOverlay(
      '✅ Mise à jour téléchargée',
      'Enregistrée dans Téléchargements : ' + result.filename +
      '. VoxLux s’arrête — ouvrez ce fichier pour installer la nouvelle version.'
    );
  } catch (e) {
    toast('❌ ' + e.message); btn.disabled = false; btn.textContent = 'Télécharger';
  }
});
checkForUpdate();

// ── Sauvegarde : fichier réel sur le serveur (source de vérité) + copie
//    instantanée en localStorage (filet de secours si le serveur est
//    momentanément injoignable — ex. redémarrage) ───────────────────────────
const AUTOSAVE_KEY = 'voxlux_autosave_v1';
let autosaveTimer = null;
let savingToServer = false;

function buildProjectSnapshot() {
  return {
    version: 1,
    name: projectName,
    savedAt: new Date().toISOString(),
    sourceFilename: selectedFile ? selectedFile.name : (window.__voxluxSourceName || ''),
    detectedLang, segs, words, trs
  };
}
async function saveProjectToServer() {
  if (!projectName || !segs.length || savingToServer) return;
  savingToServer = true;
  const stateEl = document.getElementById('project-save-state');
  if (stateEl) stateEl.textContent = 'Enregistrement…';
  try {
    const res = await fetch('/api/project/save', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ name: projectName, old_filename: projectFilename, project: buildProjectSnapshot() })
    });
    const data = await res.json();
    if (data.filename) projectFilename = data.filename;
    if (data.name && data.name !== projectName) { projectName = data.name; document.getElementById('project-name-input').value = projectName; }
    projectSavedOnce = true;
    markProjectSaved();
  } catch (e) {
    if (stateEl) stateEl.textContent = '⚠️ Sauvegarde serveur indisponible (copie locale conservée)';
  } finally {
    savingToServer = false;
  }
}
function autosaveNow() {
  if (!segs.length) return;
  try { localStorage.setItem(AUTOSAVE_KEY, JSON.stringify(buildProjectSnapshot())); }
  catch (e) { /* quota dépassé / stockage bloqué — silencieux, ce n'est qu'un filet de sécurité */ }
  // L'autosave silencieuse n'écrit le fichier réel qu'APRÈS le premier
  // enregistrement explicite (comme After Effects) — avant ça, seul le
  // filet localStorage protège le travail.
  if (projectSavedOnce) saveProjectToServer();
}
function scheduleAutosave() {
  clearTimeout(autosaveTimer);
  autosaveTimer = setTimeout(autosaveNow, 1500);
}
function clearAutosave() {
  try { localStorage.removeItem(AUTOSAVE_KEY); } catch (e) {}
}

// Au chargement de la page : proposer de restaurer une session non enregistrée
// sur le serveur (ex. la sauvegarde serveur avait échoué juste avant un
// plantage — le filet localStorage, lui, a quand même capté la dernière frappe).
(function restoreAutosaveIfAny() {
  let raw;
  try { raw = localStorage.getItem(AUTOSAVE_KEY); } catch (e) { return; }
  if (!raw) return;
  let data;
  try { data = JSON.parse(raw); } catch (e) { return; }
  if (!data || !data.segs || !data.segs.length) return;
  const when = data.savedAt ? new Date(data.savedAt).toLocaleString('fr-FR') : '';
  const ok = confirm(
    'Une session non enregistrée a été retrouvée (' + data.segs.length + ' segments' +
    (when ? ', sauvegardée automatiquement le ' + when : '') + ').\\n\\n' +
    'Reprendre cette session ?'
  );
  if (!ok) { clearAutosave(); return; }
  applyProjectData(data, '♻️ Session restaurée automatiquement — glissez la vidéo pour continuer');
})();

// ── Transcription ─────────────────────────────────────────────────────────
document.getElementById('tbtn').addEventListener('click', async () => {
  const file = selectedFile; if (!file) return;
  if (!resumeMode && segs.length > 0) {
    if (!confirm("Une nouvelle transcription effacera la transcription en cours. Continuer ?")) return;
    segs = []; words = []; trs = {};
    clearAutosave();
    projectName = ''; projectFilename = ''; projectSavedOnce = false;
    document.getElementById('project-bar').style.display = 'none';
    document.getElementById('project-bar').classList.remove('unsaved');
  }
  document.getElementById('tbtn').disabled = true;
  goTo(2);
  status('Envoi du fichier…', true);

  const form = new FormData(); form.append('file', file);
  let data;
  try {
    const res = await fetch('/api/upload', { method: 'POST', body: form });
    data = await res.json();
  } catch(e) { status('❌ Erreur upload : ' + e.message); return; }
  if (data.error) { status('❌ ' + data.error); return; }
  fileId = data.file_id;

  setupPlayer(data.file_id, data.ext);

  // Mode reprise : la vidéo ne sert qu'à la lecture, pas de nouvelle transcription
  if (resumeMode) {
    renderTable();
    renderLangCards();
    document.getElementById('master-export-btn').disabled = false;
    document.getElementById('goto3-btn').disabled = false;
    document.getElementById('addbtn').disabled = false;
    document.getElementById('smart-toggle-btn').disabled  = (words.length === 0);
    document.getElementById('script-toggle-btn').disabled = (words.length === 0);
    document.getElementById('si2').classList.add('active');
    if (detectedLang) {
      document.getElementById('si2-label').textContent = 'Transcription · ' + detectedLang;
      const lbl = document.getElementById('export-lang-label');
      if (lbl) lbl.textContent = detectedLang;
      currentLangue = detectedLang.toLowerCase();
    }
    document.getElementById('tbtn').disabled = false;
    toast('✅ Projet repris — ' + segs.length + ' segments');
    return;
  }

  status('Transcription en cours… (quelques minutes)', true);

  currentLangue  = document.getElementById('slang').value;
  currentLexicon = document.getElementById('lexicon-input').value.trim();
  try {
    const res = await fetch('/api/transcribe', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({
        file_id: fileId,
        langue:  currentLangue,
        lexicon: currentLexicon
      })
    });
    const td = await res.json();
    if (td.error) { status('❌ ' + td.error); return; }
    segs  = td.segments;
    words = td.words || [];
    renderTable();
    document.getElementById('master-export-btn').disabled = false;
    document.getElementById('goto3-btn').disabled = false;
    document.getElementById('addbtn').disabled = false;
    document.getElementById('smart-toggle-btn').disabled  = (words.length === 0);
    document.getElementById('script-toggle-btn').disabled = (words.length === 0);
    document.getElementById('si2').classList.add('active');
    const detLang = td.detected_language || '';
    detectedLang = detLang.toUpperCase() || 'FR';
    if (detLang) {
      document.getElementById('si2-label').textContent = 'Transcription · ' + detLang.toUpperCase();
      const lbl = document.getElementById('export-lang-label');
      if (lbl) lbl.textContent = detLang.toUpperCase();
    }
    // Comme After Effects/Premiere : à la fin de la transcription, on propose
    // explicitement d'enregistrer le projet (nom éditable, emplacement annoncé).
    // La barre projet est affichée dès maintenant (état "non enregistré") pour
    // qu'elle reste visible même si l'utilisateur choisit "Plus tard".
    const defaultProjectName = file.name.replace(/\.[^.]+$/, '');
    setProjectName(defaultProjectName);
    markProjectUnsaved();
    openSavePrompt(defaultProjectName);
    status('✅ ' + segs.length + ' segments transcrits' + (detLang ? ' — langue : ' + detLang : ''));
  } catch(e) {
    status('❌ ' + e.message);
  } finally {
    document.getElementById('tbtn').disabled = false;
  }
});

// ── Lecteur ───────────────────────────────────────────────────────────────
function setupPlayer(fid, ext) {
  const wrap = document.getElementById('vwrap');
  const url  = '/api/media/' + fid;
  const audioExts = new Set(['.mp3','.wav','.m4a','.ogg','.flac']);
  const navHint =
    '<div class="nav-hint">' +
      '<div class="nav-hint-item"><span class="nav-key">espace</span> lecture / pause</div>' +
      '<div class="nav-hint-item"><span class="nav-key">←</span><span class="nav-key">→</span> ±0,5 s</div>' +
      '<div class="nav-hint-item"><span class="nav-key">⇧</span> ±5 s &nbsp;·&nbsp; <span class="nav-key">alt</span> ±0,1 s</div>' +
    '</div>';

  if (audioExts.has(ext)) {
    wrap.innerHTML =
      '<div class="audio-wrap">' +
      '<audio id="mplayer" controls src="' + url + '"></audio>' +
      '<div class="audio-sub" id="suboverlay"></div>' +
      '</div>' + navHint;
  } else {
    wrap.innerHTML =
      '<video id="mplayer" controls src="' + url + '"></video>' +
      '<div class="sub-overlay" id="suboverlay"></div>' +
      '<div class="tc-disp" id="tcdisp">00:00.000</div>' + navHint;
  }
  media = document.getElementById('mplayer');
  media.addEventListener('timeupdate', onTime);
}

// ── Overlay sous-titres ───────────────────────────────────────────────────
function onTime() {
  if (!media) return;
  const t = media.currentTime;
  const tc = document.getElementById('tcdisp');
  if (tc) tc.textContent = fmtTC(t);
  const overlay = document.getElementById('suboverlay');
  const seg = segs.find(s => t >= s.start && t < s.end);
  if (overlay) overlay.textContent = seg ? seg.text : '';
  document.querySelectorAll('#stbody tr').forEach(tr => {
    const i = parseInt(tr.dataset.idx);
    const active = segs[i] && t >= segs[i].start && t < segs[i].end;
    tr.classList.toggle('active', active);
    if (active) tr.scrollIntoView({ block: 'nearest', behavior: 'smooth' });
  });
}

// ── Rendu tableau ─────────────────────────────────────────────────────────
function renderTable() {
  const tbody = document.getElementById('stbody');
  const tbl   = document.getElementById('stbl');
  const hint  = document.getElementById('ehint');
  if (!segs.length) {
    tbl.style.display = 'none'; hint.style.display = '';
    document.getElementById('segcount').textContent = '';
    return;
  }
  tbl.style.display = ''; hint.style.display = 'none';
  document.getElementById('segcount').textContent = segs.length + ' segments';
  document.getElementById('addbtn').disabled = false;
  tbody.innerHTML = '';
  segs.forEach((seg, idx) => addRow(seg, idx, tbody));
  // Relance autoResize après que le DOM soit peint
  requestAnimationFrame(() => {
    document.querySelectorAll('#stbody textarea.tt').forEach(ta => autoResize(ta));
  });
  scheduleAutosave();
}

// Ligne actuellement "active" (dernier champ cliqué / focus) — sert à savoir
// où insérer une nouvelle ligne avec "+ Ajouter".
let activeRowIndex = null;
document.getElementById('stbody').addEventListener('focusin', e => {
  const tr = e.target.closest('tr[data-idx]');
  if (tr) activeRowIndex = parseInt(tr.dataset.idx);
});

function addRow(seg, idx, tbody) {
  const tr = document.createElement('tr');
  tr.dataset.idx = idx;
  const warn = seg.text.length > 84;
  tr.innerHTML =
    '<td class="tnum">' + (idx+1) + '</td>' +
    '<td class="ttime"><input class="ti" data-f="start" value="' + fmtT(seg.start) + '"></td>' +
    '<td class="ttime"><input class="ti" data-f="end"   value="' + fmtT(seg.end)   + '"></td>' +
    '<td><textarea class="tt' + (warn?' warn':'') + '">' + esc(seg.text) + '</textarea></td>' +
    '<td class="tdel">' +
      '<button class="rebtn" title="Retranscrire ce passage">&#127908;</button>' +
      '<button class="delbtn" title="Supprimer">✕</button>' +
    '</td>';

  tr.addEventListener('click', e => {
    activeRowIndex = parseInt(tr.dataset.idx);
    if (['INPUT','TEXTAREA','BUTTON'].includes(e.target.tagName)) return;
    if (media) media.currentTime = segs[parseInt(tr.dataset.idx)].start;
  });

  tr.querySelectorAll('.ti').forEach(inp => {
    inp.addEventListener('click', e => e.stopPropagation());
    inp.addEventListener('blur', () => {
      const i = parseInt(tr.dataset.idx);
      const v = parseT(inp.value);
      if (!isNaN(v) && v >= 0) {
        segs[i][inp.dataset.f] = r1(v); inp.value = fmtT(segs[i][inp.dataset.f]);
        cascadeTime(i, inp.dataset.f, segs[i][inp.dataset.f]);
        scheduleAutosave();
      } else { inp.value = fmtT(segs[i][inp.dataset.f]); }
    });
    inp.addEventListener('keydown', e => {
      if (e.key === 'ArrowUp' || e.key === 'ArrowDown') {
        e.preventDefault();
        const i = parseInt(tr.dataset.idx);
        const step = e.shiftKey ? 1.0 : 0.1;
        const v = r1((segs[i][inp.dataset.f] || 0) + (e.key === 'ArrowUp' ? step : -step));
        if (v >= 0) { segs[i][inp.dataset.f] = v; inp.value = fmtT(v); cascadeTime(i, inp.dataset.f, v); }
      }
    });
  });

  const ta = tr.querySelector('textarea');
  autoResize(ta);
  ta.addEventListener('click', e => {
    e.stopPropagation();
    activeRowIndex = parseInt(tr.dataset.idx);
    if (media) media.currentTime = segs[activeRowIndex].start;
  });
  ta.addEventListener('input', () => {
    const i = parseInt(tr.dataset.idx);
    segs[i].text = ta.value;
    ta.classList.toggle('warn', ta.value.length > 84);
    autoResize(ta);
    scheduleAutosave();
  });
  ta.addEventListener('keydown', e => {
    if (e.key !== 'Enter') return;
    e.preventDefault();
    const i = parseInt(tr.dataset.idx);
    const pos = ta.selectionStart, full = segs[i].text;
    const t1 = full.substring(0, pos).trim(), t2 = full.substring(pos).trim();
    if (!t1 || !t2) return;
    const ratio = pos / full.length;
    const midT  = r1(segs[i].start + (segs[i].end - segs[i].start) * ratio);
    segs.splice(i, 1,
      { ...segs[i], text: t1, end: midT },
      { id: i + 0.5, start: midT, end: segs[i] ? segs[i].end : midT + 3, text: t2 }
    );
    segs.forEach((s, idx) => { s.id = idx; });
    renderTable();
    setTimeout(() => {
      const rows = document.querySelectorAll('#stbody tr');
      if (rows[i+1]) { const nt = rows[i+1].querySelector('textarea'); nt.focus(); nt.setSelectionRange(0,0); }
    }, 0);
  });

  tr.querySelector('.delbtn').addEventListener('click', e => {
    e.stopPropagation();
    const i = parseInt(tr.dataset.idx);
    const deleted = segs[i];
    // Si le champ est vide (texte coupé-collé ailleurs) et qu'une ligne
    // précédente existe, on reporte son timecode de fin sur cette ligne
    // précédente — ça correspond souvent à la réalité du montage.
    const wasEmptied = !deleted.text.trim() && i > 0;
    segs.splice(i, 1);
    if (wasEmptied && segs[i - 1]) {
      segs[i - 1].end = deleted.end;
      cascadeTime(i - 1, 'end', deleted.end);
      toast('Fin de la ligne précédente ajustée à ' + fmtT(deleted.end));
    }
    renderTable();
  });

  tr.querySelector('.rebtn').addEventListener('click', async e => {
    e.stopPropagation();
    const i = parseInt(tr.dataset.idx);
    const seg = segs[i];
    if (!fileId) { toast('❌ Aucune vidéo chargée — glissez le fichier d’origine'); return; }
    const btn = e.currentTarget;
    const original = btn.innerHTML;
    btn.disabled = true; btn.innerHTML = '&#8987;';
    try {
      const res = await fetch('/api/retranscribe_segment', {
        method: 'POST', headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({
          file_id: fileId, start: seg.start, end: seg.end,
          langue: currentLangue, lexicon: currentLexicon
        })
      });
      const data = await res.json();
      if (data.error) { toast('❌ ' + data.error); return; }
      if (!data.text) { toast('⚠️ Toujours rien détecté sur ce passage'); return; }
      segs[i].text = data.text;
      renderTable();
      toast('🎙 Passage retranscrit');
    } catch (err) {
      toast('❌ ' + err.message);
    } finally {
      btn.disabled = false; btn.innerHTML = original;
    }
  });

  tbody.appendChild(tr);
}

function autoResize(ta) { ta.style.height = 'auto'; ta.style.height = ta.scrollHeight + 'px'; }

// ── Ajouter ligne ─────────────────────────────────────────────────────────
document.getElementById('addbtn').addEventListener('click', () => {
  // Insère juste après la ligne active (dernier champ cliqué/édité) ;
  // à défaut (aucune ligne active), ajoute à la fin comme avant.
  const insertAt = (activeRowIndex !== null && activeRowIndex >= 0 && activeRowIndex < segs.length)
    ? activeRowIndex + 1 : segs.length;
  const prev = segs[insertAt - 1];
  const next = segs[insertAt];
  let start, end;
  if (prev) {
    start = r1(prev.end);
    end = next ? r1(Math.min(prev.end + 3, next.start)) : r1(prev.end + 3);
    if (end <= start) end = r1(start + 0.5); // filet si les segments voisins sont très serrés
  } else {
    start = 0;
    end = next ? r1(Math.min(3, next.start)) : 3;
  }
  segs.splice(insertAt, 0, { id: insertAt, start, end, text: '' });
  segs.forEach((s, idx) => { s.id = idx; });
  activeRowIndex = insertAt;
  renderTable();
  const rows = document.querySelectorAll('#stbody tr');
  if (rows[insertAt]) rows[insertAt].querySelector('textarea').focus();
});

// ── Sync depuis tableau ───────────────────────────────────────────────────
function syncFromTable() {
  document.querySelectorAll('#stbody tr').forEach(tr => {
    const i = parseInt(tr.dataset.idx);
    if (i < 0 || i >= segs.length) return;
    const si = tr.querySelector('[data-f="start"]');
    const ei = tr.querySelector('[data-f="end"]');
    const ti = tr.querySelector('textarea');
    if (si) { const v = parseT(si.value); if (!isNaN(v)) segs[i].start = r1(v); }
    if (ei) { const v = parseT(ei.value); if (!isNaN(v)) segs[i].end   = r1(v); }
    if (ti) segs[i].text = ti.value;
  });
}

// ── Export master ─────────────────────────────────────────────────────────
document.getElementById('master-export-btn').addEventListener('click', async () => {
  syncFromTable();
  const fmts = [];
  if (document.getElementById('m-xlsx').checked) fmts.push('xlsx');
  if (document.getElementById('m-srt').checked)  fmts.push('srt');
  if (document.getElementById('m-vtt').checked)  fmts.push('vtt');
  if (!fmts.length) { status('Sélectionnez au moins un format'); return; }
  status('Export en cours…', true);
  await exportSegments(segs, fmts, 'transcription-' + (detectedLang || 'fr'));
  status('✅ Transcription téléchargée');
});

// ── Vers étape 3 ──────────────────────────────────────────────────────────
document.getElementById('goto3-btn').addEventListener('click', () => {
  syncFromTable();
  document.getElementById('si3').classList.remove('done');
  goTo(3);
  document.getElementById('translate-all-btn').disabled = (Object.keys(trs).length === 0);
});

// ── Étape 3 : ajouter langue ──────────────────────────────────────────────
document.getElementById('add-lang-btn').addEventListener('click', () => {
  const lang = document.getElementById('tlang-pick').value;
  if (trs[lang]) { status('Langue déjà ajoutée'); return; }
  trs[lang] = { status: 'wait', segments: [] };
  renderLangCards();
  document.getElementById('translate-all-btn').disabled = false;
});

function renderLangCards() {
  const container = document.getElementById('lang-cards');
  const langs = Object.keys(trs);
  if (!langs.length) {
    container.innerHTML = '<div class="lang-card-empty" id="lang-empty">Ajoutez des langues cibles ci-dessus pour commencer.</div>';
    document.getElementById('translate-all-btn').disabled = true;
    return;
  }
  const openPreviews = new Set();
  container.querySelectorAll('.lang-card').forEach(c => {
    if (c.querySelector('.lang-preview.open')) openPreviews.add(c.dataset.lang);
  });
  container.innerHTML = '';
  langs.forEach(lang => {
    const state = trs[lang];
    const isDone = state.status === 'done';
    const isRunning = state.status === 'running';
    const badgeClass = { wait:'badge-wait', running:'badge-running', done:'badge-done', error:'badge-error' }[state.status];
    const badgeLabel = { wait:'En attente', running:'Traduction…', done:'✅ Terminé', error:'❌ Erreur' }[state.status];
    const previewOpen = openPreviews.has(lang);

    const card = document.createElement('div');
    card.className = 'lang-card'; card.dataset.lang = lang;

    const header = document.createElement('div');
    header.className = 'lang-card-header';
    header.innerHTML =
      '<span class="lang-name">' + (LANG_NAMES[lang] || lang) + '</span>' +
      '<span class="lang-badge ' + badgeClass + '">' + (isRunning ? '<span class="spin"></span> ' : '') + badgeLabel + '</span>' +
      '<div class="lang-exports">' +
        '<button class="btn btn-green btn-sm exp-xlsx-btn" ' + (isDone?'':'disabled') + '>Excel</button>' +
        '<button class="btn btn-green btn-sm exp-srt-btn"  ' + (isDone?'':'disabled') + '>SRT</button>' +
        '<button class="btn btn-green btn-sm exp-vtt-btn"  ' + (isDone?'':'disabled') + '>WebVTT</button>' +
      '</div>' +
      (isDone ? '<span class="lang-expand ' + (previewOpen?'open':'') + '">▼</span>' : '<span style="width:20px"></span>') +
      '<button class="lang-remove" title="Retirer">✕</button>';

    const preview = document.createElement('div');
    preview.className = 'lang-preview' + (previewOpen ? ' open' : '');
    if (isDone && state.segments.length) {
      // En-tête colonnes
      const hdr = document.createElement('div');
      hdr.className = 'preview-header';
      hdr.innerHTML =
        '<span class="preview-header-tc">TC</span>' +
        '<span class="preview-header-orig">Transcription originale</span>' +
        '<span class="preview-header-tr">' + (LANG_NAMES[lang] || lang) + '</span>';
      preview.appendChild(hdr);
      // Lignes : original (segs[i]) + traduction (state.segments[i])
      state.segments.forEach((seg, i) => {
        const origSeg = segs[i];
        const row = document.createElement('div');
        row.className = 'preview-seg';
        const tcSpan = document.createElement('span');
        tcSpan.className = 'preview-tc';
        tcSpan.textContent = fmtT(seg.start);
        const origSpan = document.createElement('span');
        origSpan.className = 'preview-orig';
        origSpan.textContent = origSeg ? origSeg.text : '';
        const trDiv = document.createElement('div');
        trDiv.className = 'preview-text';
        trDiv.contentEditable = 'true';
        trDiv.spellcheck = false;
        trDiv.textContent = seg.text;
        trDiv.addEventListener('input', () => {
          trs[lang].segments[i].text = trDiv.textContent;
          scheduleAutosave();
        });
        trDiv.addEventListener('keydown', e => e.stopPropagation());
        row.appendChild(tcSpan);
        row.appendChild(origSpan);
        row.appendChild(trDiv);
        preview.appendChild(row);
      });
      const hint = document.createElement('div');
      hint.className = 'preview-edit-hint';
      hint.textContent = "Cliquer pour éditer une traduction";
      preview.appendChild(hint);
    }

    header.addEventListener('click', e => {
      if (e.target.classList.contains('lang-remove') || e.target.closest('.lang-exports')) return;
      if (!isDone) return;
      preview.classList.toggle('open');
      const icon = header.querySelector('.lang-expand');
      if (icon) icon.classList.toggle('open');
    });

    header.querySelector('.exp-xlsx-btn').addEventListener('click', e => { e.stopPropagation(); exportSegments(state.segments, ['xlsx'], lang); });
    header.querySelector('.exp-srt-btn') .addEventListener('click', e => { e.stopPropagation(); exportSegments(state.segments, ['srt'],  lang); });
    header.querySelector('.exp-vtt-btn') .addEventListener('click', e => { e.stopPropagation(); exportSegments(state.segments, ['vtt'],  lang); });
    header.querySelector('.lang-remove') .addEventListener('click', e => {
      e.stopPropagation(); delete trs[lang]; renderLangCards();
    });

    card.appendChild(header); card.appendChild(preview);
    container.appendChild(card);
  });

  document.getElementById('translate-all-btn').disabled = (langs.length === 0);
  const anyDone = Object.values(trs).some(t => t.status === 'done');
  document.getElementById('export-all-btn').disabled = !anyDone;
  document.getElementById('p3-footer').classList.toggle('visible', anyDone);
  scheduleAutosave();
}

// ── Tout traduire ─────────────────────────────────────────────────────────
document.getElementById('translate-all-btn').addEventListener('click', async () => {
  const langs = Object.keys(trs).filter(l => trs[l].status !== 'done');
  if (!langs.length) { status('Toutes les langues sont déjà traduites'); return; }
  document.getElementById('translate-all-btn').disabled = true;
  for (const lang of langs) {
    trs[lang].status = 'running'; renderLangCards();
    status('Traduction ' + (LANG_NAMES[lang]||lang) + '…', true);
    try {
      const res = await fetch('/api/translate', {
        method: 'POST', headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ segments: segs, target_lang: lang })
      });
      const data = await res.json();
      if (data.error) throw new Error(data.error);
      trs[lang].segments = data.segments; trs[lang].status = 'done';
    } catch(e) {
      trs[lang].status = 'error';
      status('❌ ' + (LANG_NAMES[lang]||lang) + ' : ' + e.message);
    }
    renderLangCards();
  }
  document.getElementById('translate-all-btn').disabled = false;
  document.getElementById('export-all-btn').disabled = false;
  status('✅ Traductions terminées');
});

// ── Tout exporter ─────────────────────────────────────────────────────────
document.getElementById('export-all-btn').addEventListener('click', async () => {
  const fmts = [];
  if (document.getElementById('e-xlsx').checked) fmts.push('xlsx');
  if (document.getElementById('e-srt').checked)  fmts.push('srt');
  if (document.getElementById('e-vtt').checked)  fmts.push('vtt');
  if (!fmts.length) { status('Sélectionnez au moins un format'); return; }
  const done = Object.entries(trs).filter(([,v]) => v.status === 'done');
  if (!done.length) { status('Aucune traduction terminée'); return; }
  status('Export en cours…', true);
  for (const [lang, state] of done) {
    await exportSegments(state.segments, fmts, lang);
    await new Promise(r => setTimeout(r, 200));
  }
  status('✅ Tous les fichiers exportés');
});

// ── Navigation clavier vidéo ──────────────────────────────────────────────
document.addEventListener('keydown', e => {
  if (!media) return;
  const tag = document.activeElement ? document.activeElement.tagName : '';
  if (tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT') return;
  if (!document.getElementById('p2').classList.contains('active')) return;
  // Barre d'espace : play / pause
  if (e.key === ' ') {
    e.preventDefault();
    media.paused ? media.play() : media.pause();
    return;
  }
  let delta = 0;
  if (e.key === 'ArrowLeft')  delta = e.shiftKey ? -5 : e.altKey ? -0.1 : -0.5;
  if (e.key === 'ArrowRight') delta = e.shiftKey ?  5 : e.altKey ?  0.1 :  0.5;
  if (!delta) return;
  e.preventDefault();
  media.currentTime = Math.max(0, Math.min(media.duration || Infinity, media.currentTime + delta));
  const tc = document.getElementById('tcdisp');
  if (tc) {
    tc.style.background = 'rgba(167,139,250,0.5)';
    setTimeout(() => { if (tc) tc.style.background = 'rgba(0,0,0,0.5)'; }, 200);
  }
});

// ── Cascade timecodes ─────────────────────────────────────────────────────
function cascadeTime(idx, field, newVal) {
  if (field === 'end' && idx + 1 < segs.length) {
    if (newVal > segs[idx + 1].start) {
      segs[idx + 1].start = newVal;
      const nextRow = document.querySelector('#stbody tr[data-idx="' + (idx+1) + '"]');
      if (nextRow) { const inp = nextRow.querySelector('[data-f="start"]'); if (inp) inp.value = fmtT(newVal); }
    }
  }
  if (field === 'start' && idx > 0) {
    if (newVal < segs[idx - 1].end) {
      segs[idx - 1].end = newVal;
      const prevRow = document.querySelector('#stbody tr[data-idx="' + (idx-1) + '"]');
      if (prevRow) { const inp = prevRow.querySelector('[data-f="end"]'); if (inp) inp.value = fmtT(newVal); }
    }
  }
}

// ── Découpage intelligent ─────────────────────────────────────────────────
document.getElementById('smart-toggle-btn').addEventListener('click', () => {
  document.getElementById('smart-bar').classList.toggle('open');
  document.getElementById('script-bar').classList.remove('open');
});
document.getElementById('sr-pause').addEventListener('input', function() {
  document.getElementById('sr-pause-val').textContent = parseFloat(this.value).toFixed(2) + ' s';
});
document.getElementById('sr-chars').addEventListener('input', function() {
  document.getElementById('sr-chars-val').textContent = this.value + ' car.';
});
document.getElementById('sr-dur').addEventListener('input', function() {
  document.getElementById('sr-dur-val').textContent = parseFloat(this.value).toFixed(1) + ' s';
});
document.getElementById('smart-apply-btn').addEventListener('click', async () => {
  if (!words.length) { status('Aucune donnée de mots — relancez la transcription'); return; }
  const minPause    = parseFloat(document.getElementById('sr-pause').value);
  const maxChars    = parseInt(document.getElementById('sr-chars').value);
  const minDuration = parseFloat(document.getElementById('sr-dur').value);
  status('Découpage en cours…', true);
  try {
    const res = await fetch('/api/smart_split', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ words, min_pause: minPause, max_chars: maxChars, min_duration: minDuration })
    });
    const data = await res.json();
    if (data.error) { toast('❌ ' + data.error); return; }
    segs = data.segments; renderTable();
    document.getElementById('smart-bar').classList.remove('open');
    toast('✅ Découpage intelligent appliqué — ' + segs.length + ' segments');
  } catch(e) { toast('❌ ' + e.message); }
});

// ── Alignement script ─────────────────────────────────────────────────────
document.getElementById('script-toggle-btn').addEventListener('click', () => {
  document.getElementById('script-bar').classList.toggle('open');
  document.getElementById('smart-bar').classList.remove('open');
});
document.getElementById('align-btn').addEventListener('click', async () => {
  const script = document.getElementById('script-input').value.trim();
  if (!script) { status('Le script est vide'); return; }
  if (!words.length) { status('Aucun timestamp de mots — relancez la transcription'); return; }
  status('Alignement du script en cours…', true);
  try {
    const res = await fetch('/api/align_script', {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ script, words })
    });
    const data = await res.json();
    if (data.error) { status('❌ ' + data.error); return; }
    segs = data.segments; renderTable();
    document.getElementById('script-bar').classList.remove('open');
    status('✅ ' + segs.length + ' phrases alignées sur les timecodes vidéo');
  } catch(e) { status('❌ ' + e.message); }
});

// ── Export générique ──────────────────────────────────────────────────────
async function exportSegments(segments, fmts, prefix) {
  for (const fmt of fmts) {
    try {
      const res = await fetch('/api/export/' + fmt, {
        method: 'POST', headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ segments })
      });
      if (!res.ok) continue;
      const blob = await res.blob();
      const url  = URL.createObjectURL(blob);
      const ext  = fmt === 'xlsx' ? '.xlsx' : fmt === 'srt' ? '.srt' : '.vtt';
      const a    = document.createElement('a');
      a.href = url; a.download = prefix + ext;
      document.body.appendChild(a); a.click();
      document.body.removeChild(a);
      URL.revokeObjectURL(url);
      await new Promise(r => setTimeout(r, 150));
    } catch(e) { console.error(fmt, e); }
  }
}
</script>
</body>
</html>"""
# ─── Lancement ────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    # En mode app (.app/.bat), le lanceur ouvre le navigateur lui-même →
    # on évite la double ouverture via VOXLUX_NO_AUTOOPEN.
    if not os.environ.get('VOXLUX_NO_AUTOOPEN'):
        threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:7860')).start()
    print("✅ VoxLux V3 démarré — http://127.0.0.1:7860")
    app.run(host='127.0.0.1', port=7860, debug=False)
